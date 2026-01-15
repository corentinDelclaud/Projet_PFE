import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from pathlib import Path
import re

def create_tracking_excel(output_path):
    """Crée le fichier Excel de suivi avec tous les onglets"""
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # Style pour les en-têtes
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    
    # === ONGLET 1: Paramètres du Modèle ===
    ws1 = wb.create_sheet("Paramètres")
    ws1.append(["Paramètre", "Valeur", "Description"])
    params = [
        ["max_time_in_seconds", "", "Temps limite du solveur (s)"],
        ["num_search_workers", "", "Nombre de workers parallèles"],
        ["QUOTA_PENALTY_WEIGHT", "", "Poids de pénalité quota"],
        ["LISSAGE_THRESHOLD", "", "Seuil de lissage"],
        ["delta_max (quota < seuil)", "", "Écart max si quota bas"],
        ["delta_max (quota >= seuil)", "", "Écart max si quota élevé"],
        ["Poids préférence", "", "Bonus jour préférence"],
        ["Poids base", "", "Poids affectation standard"]
    ]
    for row in params:
        ws1.append(row)
    
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    
    # === ONGLET 2: Journal des Exécutions ===
    ws2 = wb.create_sheet("Journal_Executions")
    headers = ["Date", "Heure", "Status", "Valeur_Objectif", "Best_Bound", "Gap_Integral",
               "Temps_Total(s)", "Temps_UserTime(s)", "Branches", "Conflicts", "Propagations",
               "LP_Iterations", "Variables", "Contraintes_Bool", "Contraintes_Int", "Notes"]
    ws2.append(headers)
    
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    
    # === ONGLET 3: Métriques Importantes ===
    ws3 = wb.create_sheet("Metriques_Cles")
    ws3.append(["Métrique", "Valeur", "Calcul/Source"])
    metrics = [
        ["Status", "", "OPTIMAL/FEASIBLE/INFEASIBLE"],
        ["Valeur objectif", "", "objective dans output"],
        ["Best bound", "", "best_bound dans output"],
        ["Gap (objectif - bound)", "", "Calculé"],
        ["Gap integral", "", "gap_integral dans output"],
        ["Temps total (s)", "", "walltime"],
        ["Temps user (s)", "", "usertime"],
        ["Nombre de branches", "", "branches"],
        ["Nombre de conflits", "", "conflicts"],
        ["Propagations", "", "propagations"],
        ["Itérations LP", "", "lp_iterations"],
        ["Variables (int + bool)", "", "integers + booleans"],
        ["Solutions trouvées", "", "Solutions (count)"],
        ["LNS améliorations", "", "Somme Improv/Calls"],
        ["Taux succès LNS (%)", "", "(Améliorations / Appels) * 100"]
    ]
    for row in metrics:
        ws3.append(row)
    
    for cell in ws3[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    
    # === ONGLET 4: Analyse Résultats Planning ===
    ws4 = wb.create_sheet("Analyse_Planning")
    ws4.append(["Métrique", "Valeur", "Calcul"])
    planning_metrics = [
        ["Nombre affectations totales", "", "Comptage lignes CSV"],
        ["Élèves uniques affectés", "", "Count unique Id_Eleve"],
        ["Disciplines utilisées", "", "Count unique Discipline"],
        ["Vacations utilisées", "", "Count unique (Semaine, Jour, Periode)"],
        ["Taux remplissage quotas (%)", "", "(Affectations / Quotas totaux) * 100"],
        ["Affectations sur jour préféré", "", "Comptage"],
        ["Taux satisfaction préférences (%)", "", "(Préférences / Total) * 100"]
    ]
    for row in planning_metrics:
        ws4.append(row)
    
    for cell in ws4[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    
    # === ONGLET 5: Performance LNS ===
    ws5 = wb.create_sheet("Performance_LNS")
    ws5.append(["Stratégie LNS", "Améliorations", "Appels", "Taux_Succès(%)", "Closed(%)", "Difficulty", "TimeLimit"])
    
    for cell in ws5[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    
    # === ONGLET 6: Comparaisons Versions ===
    ws6 = wb.create_sheet("Comparaisons")
    ws6.append(["Version", "Date", "Status", "Valeur_Obj", "Gap", "Temps(s)", "Solutions", "Meilleure"])
    
    for cell in ws6[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
    
    # Ajuster largeur colonnes
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output_path)
    print(f"✅ Fichier Excel créé : {output_path}")


def parse_ortools_output(log_text):
    """Parse la sortie d'OR-Tools et extrait les métriques importantes"""
    
    metrics = {}
    
    # Status et objectif
    status_match = re.search(r'status: (\w+)', log_text)
    metrics['status'] = status_match.group(1) if status_match else "UNKNOWN"
    
    obj_match = re.search(r'objective: (-?\d+)', log_text)
    metrics['objective'] = int(obj_match.group(1)) if obj_match else None
    
    bound_match = re.search(r'best_bound: (-?\d+)', log_text)
    metrics['best_bound'] = int(bound_match.group(1)) if bound_match else None
    
    gap_match = re.search(r'gap_integral: (\d+)', log_text)
    metrics['gap_integral'] = int(gap_match.group(1)) if gap_match else None
    
    # Temps
    walltime_match = re.search(r'walltime: ([\d.]+)', log_text)
    metrics['walltime'] = float(walltime_match.group(1)) if walltime_match else None
    
    usertime_match = re.search(r'usertime: ([\d.]+)', log_text)
    metrics['usertime'] = float(usertime_match.group(1)) if usertime_match else None
    
    # Variables et contraintes
    integers_match = re.search(r'integers: (\d+)', log_text)
    metrics['integers'] = int(integers_match.group(1)) if integers_match else None
    
    booleans_match = re.search(r'booleans: (\d+)', log_text)
    metrics['booleans'] = int(booleans_match.group(1)) if booleans_match else None
    
    # Stats de recherche
    conflicts_match = re.search(r"'default_lp':\s+\d+\s+(\d+)\s+", log_text)
    metrics['conflicts'] = int(conflicts_match.group(1)) if conflicts_match else None
    
    branches_match = re.search(r'branches: (\d+)', log_text)
    metrics['branches'] = int(branches_match.group(1)) if branches_match else None
    
    propagations_match = re.search(r'propagations: (\d+)', log_text)
    metrics['propagations'] = int(propagations_match.group(1)) if propagations_match else None
    
    lp_iter_match = re.search(r'lp_iterations: (\d+)', log_text)
    metrics['lp_iterations'] = int(lp_iter_match.group(1)) if lp_iter_match else None
    
    # Nombre de solutions
    solutions_match = re.search(r'Solutions \((\d+)\)', log_text)
    metrics['num_solutions'] = int(solutions_match.group(1)) if solutions_match else 0
    
    # Parser LNS stats
    lns_stats = {}
    lns_pattern = r"'(\w+)':\s+(\d+)/(\d+)\s+(\d+)%\s+([\d.e+-]+)\s+([\d.]+)"
    for match in re.finditer(lns_pattern, log_text):
        strategy = match.group(1)
        improvements = int(match.group(2))
        calls = int(match.group(3))
        closed = int(match.group(4))
        difficulty = float(match.group(5))
        time_limit = float(match.group(6))
        
        lns_stats[strategy] = {
            'improvements': improvements,
            'calls': calls,
            'success_rate': (improvements / calls * 100) if calls > 0 else 0,
            'closed': closed,
            'difficulty': difficulty,
            'time_limit': time_limit
        }
    
    metrics['lns_stats'] = lns_stats
    
    return metrics


def add_execution_to_excel(excel_path, metrics, notes=""):
    """Ajoute une exécution au journal"""
    
    df = pd.read_excel(excel_path, sheet_name="Journal_Executions")
    
    new_row = {
        'Date': datetime.now().strftime('%Y-%m-%d'),
        'Heure': datetime.now().strftime('%H:%M:%S'),
        'Status': metrics.get('status', ''),
        'Valeur_Objectif': metrics.get('objective', ''),
        'Best_Bound': metrics.get('best_bound', ''),
        'Gap_Integral': metrics.get('gap_integral', ''),
        'Temps_Total(s)': metrics.get('walltime', ''),
        'Temps_UserTime(s)': metrics.get('usertime', ''),
        'Branches': metrics.get('branches', ''),
        'Conflicts': metrics.get('conflicts', ''),
        'Propagations': metrics.get('propagations', ''),
        'LP_Iterations': metrics.get('lp_iterations', ''),
        'Variables': (metrics.get('integers', 0) + metrics.get('booleans', 0)),
        'Contraintes_Bool': metrics.get('booleans', ''),
        'Contraintes_Int': metrics.get('integers', ''),
        'Notes': notes
    }
    
    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    
    with pd.ExcelWriter(excel_path, mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name='Journal_Executions', index=False)
    
    # Ajouter aussi les stats LNS
    if 'lns_stats' in metrics and metrics['lns_stats']:
        lns_df = pd.read_excel(excel_path, sheet_name="Performance_LNS")
        
        for strategy, stats in metrics['lns_stats'].items():
            lns_row = {
                'Stratégie LNS': strategy,
                'Améliorations': stats['improvements'],
                'Appels': stats['calls'],
                'Taux_Succès(%)': round(stats['success_rate'], 2),
                'Closed(%)': stats['closed'],
                'Difficulty': stats['difficulty'],
                'TimeLimit': stats['time_limit']
            }
            lns_df = pd.concat([lns_df, pd.DataFrame([lns_row])], ignore_index=True)
        
        with pd.ExcelWriter(excel_path, mode='a', if_sheet_exists='replace') as writer:
            lns_df.to_excel(writer, sheet_name='Performance_LNS', index=False)
    
    print(f"✅ Exécution ajoutée à {excel_path}")


if __name__ == "__main__":
    # Chemins
    output_excel = Path(__file__).parent.parent.parent / "resultat" / "suivi_experimentations.xlsx"
    
    # Créer l'Excel
    create_tracking_excel(output_excel)
    
    # Exemple d'utilisation avec votre log
    log_example = """
    [Votre log OR-Tools ici]
    """
    
    # Parser et ajouter
    # metrics = parse_ortools_output(log_example)
    # add_execution_to_excel(output_excel, metrics, notes="Première exécution test")