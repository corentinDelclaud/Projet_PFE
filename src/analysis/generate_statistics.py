import sys
import os
import csv
import collections
import argparse
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import json

# Add the parent parent directory to sys.path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..')))

from src.classes.discipline import discipline
from src.classes.eleve import eleve
from src.classes.enum.niveaux import niveau
from src.classes.jour_preference import jour_pref

def load_data():
    # --- 1. Load Disciplines (Copied from model configuration) ---
    # Note: Use the same configuration as the model
    poly = discipline(1, "Polyclinique", [20,20,20,20,20,20,20,20,20,20], True, [50,50,50], [True]*10,[4,5,6])
    paro = discipline(2, "Parodontologie", [0,4,4,4,4,4,4,4,4,4], False, [6,6,6], [False, True, True, True, True, True, True, True, True, True],[4,5,6])
    como = discipline(3, "Comodulation", [3,3,3,3,3,3,0,0,3,3], False, [6,6,6] , [True, True, True, True, True, True, False, False, True, True],[4,5,6])
    pedo_soins = discipline(4, "Pédodontie Soins", [10,0,0,0,20,20,20,0,20,0], True, [12,12,12], [True]*10,[4,5,6])
    odf = discipline(5, "Orthodontie", [3] * 10, False, [4,4,4], [True]*10,[5])
    occl = discipline(6, "Occlusodontie", [4,0,4,0,0,0,4,0,4,0], False, [3,0,0], [True, False, True, False, False, False, True, False, True, False],[4])
    ra = discipline(7, "Radiologie", [4] * 10, False, [9,8,0], [True]*10,[4,5])
    ste = discipline(8, "Stérilisation", [1] * 10, False, [3,2,0], [True]*10,[4,5,6])
    pano = discipline(9, "Panoramique", [0,1,0,1,0,1,0,1,0,1,0,1] * 10, False, [0,3,3], [True]*10,[5,6])
    urg = discipline(10, "Urgence", [12] * 10, False, [20,20,20], [True]*10,[4,5,6])
    pedo_urg = discipline(11, "Pédodontie Urgences", [2] * 10, False, [0,1,1], [True]*10,[5,6])
    bloc = discipline(12, "BLOC", [0,0,2,0,0,0,2,0,0,0], False, [0,1,1], [False,False,True,False,False,False,True,False,False,False],[5,6])
    sp = discipline(13, "Soins spécifiques", [1,0,0,0,0,0,0,0,0,0], False, [0,0,0], [True,False,False,False,False,False,False,False,False,False],[5,6])   

    # --- Apply Constraints Modifications (Crucial for Stats) ---
    # Copied from model_V4_01.py lines ~55-80
    
    # Configuration des présences (Ouvertures/Fermetures créneaux types)
    poly.multiple_modif_presence([0,1,2,3,4,5,6,7,8,9], [True, True, True, True, True, True, True, True, True, True])
    paro.multiple_modif_presence([0,1,2,3,4,5,6,7,8,9], [False, True, True, True, True, True, True, True, True, True])
    como.multiple_modif_presence([0,1,2,3,4,5,6,7,8,9], [True, True, True, True, True, True, False, False, True, True])
    pedo_soins.multiple_modif_presence([0,1,2,3,4,5,6,7,8,9], [True, False, False, False, True, True, True, False, True, False])
    odf.multiple_modif_presence([0,1,2,3,4,5,6,7,8,9], [False, True, False, True, True, True, False, False, False, True])
    occl.multiple_modif_presence([0,1,2,3,4,5,6,7,8,9], [True, False, True, False, False, False, True, False, True, False])
    ra.multiple_modif_presence([0,1,2,3,4,5,6,7,8,9], [True, True, True, True, True, True, True, True, True, True])
    ste.multiple_modif_presence([0,1,2,3,4,5,6,7,8,9], [True, True, True, True, True, True, True, True, True, True])
    pano.multiple_modif_presence([0,1,2,3,4,5,6,7,8,9], [False, True, False, True, False, True, False, True, False, True])
    urg.multiple_modif_presence([0,1,2,3,4,5,6,7,8,9], [True, True, True, True, True, True, True, True, True, True])
    pedo_urg.multiple_modif_presence([0,1,2,3,4,5,6,7,8,9], [False, True, False, False, False, False, False, True, False, True])
    bloc.multiple_modif_presence([0,1,2,3,4,5,6,7,8,9], [True,False,True,False,False,False,True,False,False,False])

    # Modifications spécifiques (Contraintes avancées)
    poly.modif_nb_vacations_par_semaine(4)
    poly.modif_paire_jours([ (0,1), (2,3), (0,4) ]) # Paires requises
    poly.modif_take_jour_pref(True)
    bloc.modif_fill_requirement(True)
    occl.modif_repetition_continuite(1, 12)  # Pas 2 fois de suite
    sp.modif_fill_requirement(True)
    ste.modif_fill_requirement(True)
    ra.modif_priorite_niveau([4,5,6])
    odf.modif_repartition_semestrielle([2,2])  # 4 total, 2 in each semester
    pedo_soins.modif_frequence_vacations(2)  # Une semaine sur deux
    pedo_soins.modif_priorite_niveau([5,6,4])  # 1st prio 5A, 2nd prio 6A, 3rd prio 4A
    paro.modif_mixite_groupes(2) # au moins 2 niveaux différents par vacation
    como.modif_mixite_groupes(3) # un élève de chaque niveau
    pano.modif_priorite_niveau([6,5])
    urg.modif_remplacement_niveau([(5,6,7),(5,4,5)])
    pedo_urg.modif_priorite_niveau([6,5])

    disciplines = [poly, paro, como, pedo_soins, odf, occl, ra, ste, pano, urg, pedo_urg, bloc, sp]
    disc_map = {d.nom_discipline: d for d in disciplines}

    # --- 2. Load Students ---
    eleves = {}
    eleves_csv = os.path.join(os.path.dirname(__file__), '..', '..', 'data', 'eleves_with_code.csv')
    try:
        with open(eleves_csv, mode='r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                try:
                    eid = int(row["id_eleve"])
                    annee_val = niveau[row["annee"]].value
                    eleves[eid] = {
                        "id": eid,
                        "annee": annee_val,
                        "nom_annee": row["annee"],
                        "id_binome": int(row["id_binome"]) if row["id_binome"] else 0,
                        "jour_preference": row.get("jour_preference", "").strip().capitalize(),
                        "periode_stage": int(row["periode_stage"]) if row.get("periode_stage") else 0
                    }
                except KeyError: continue
    except FileNotFoundError:
        print(f"Error: {eleves_csv} not found.")
        return None, None
    
    # --- 3. Load Stages ---
    stages = {} # Key: (nom_annee, periode_id) -> (start_week, end_week)
    stages_csv = os.path.join(os.path.dirname(__file__), '..', '..', 'data', 'stages.csv')
    try:
        with open(stages_csv, mode='r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                if row.get("deb_semaine") and row.get("fin_semaine"):
                    try:
                        lvl = row["pour_niveau"]
                        pid = int(row["periode"])
                        start = float(row["deb_semaine"])
                        end = float(row["fin_semaine"])
                        stages[(lvl, pid)] = (start, end)
                    except ValueError:
                        continue
    except FileNotFoundError:
        print(f"Error: {stages_csv} not found.")
    
    return disc_map, eleves, stages

def analyze_solution(solution_path, disc_map, eleves, stages):
    if not os.path.exists(solution_path):
        print(f"Error: {solution_path} not found.")
        return

    # Data Structure: { student_id: { discipline_name: count } }
    assignments = collections.defaultdict(lambda: collections.defaultdict(int))
    detailed_assignments = set() # (Semaine, Jour, Apres-Midi, Discipline, Id_Eleve)

    print(f"Reading solution from {solution_path}...")
    with open(solution_path, mode='r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            if "STAGE:" in row["Discipline"]: continue
            
            try:
                sid = int(row["Id_Eleve"])
                disc_name = row["Discipline"]
                assignments[sid][disc_name] += 1
                
                # Check headers for detail extraction
                if "Semaine" in row and "Jour" in row and "Apres-Midi" in row:
                    detailed_assignments.add((int(row["Semaine"]), row["Jour"], row["Apres-Midi"], disc_name, sid))

            except ValueError: continue

    # --- Binome Analysis ---
    binome_stats = {
        "Total_Binome_Vacations_Checked": 0,
        "Respect_Binome": 0,
        "Violations": 0,
        "Percentage": 0
    }
    
    binome_disciplines = {name for name, d in disc_map.items() if d.en_binome}
    
    for (sem, jour, am, disc_name, sid) in detailed_assignments:
        if disc_name in binome_disciplines:
            if sid not in eleves: continue
            
            bin_id = eleves[sid].get("id_binome", 0)
            
            # Check if valid binome exists (and is in the list)
            if bin_id > 0 and bin_id in eleves:
                binome_stats["Total_Binome_Vacations_Checked"] += 1
                
                # Check if binome is assigned to same slot
                if (sem, jour, am, disc_name, bin_id) in detailed_assignments:
                    binome_stats["Respect_Binome"] += 1
                else:
                    binome_stats["Violations"] += 1
    
    if binome_stats["Total_Binome_Vacations_Checked"] > 0:
        binome_stats["Percentage"] = round(binome_stats["Respect_Binome"] / binome_stats["Total_Binome_Vacations_Checked"] * 100, 1)

    # --- Constraint Analysis (New) ---
    constraint_stats = []

    # Convert detailed assignments to DataFrame for easier analysis
    # Columns: Semaine, Jour, Apres-Midi, Discipline, Id_Eleve
    # detailed_assignments is a set of tuples
    
    # We need to reconstruct the "HalfDay" integer/boolean for sorting if needed, 
    # but Semaine is key for most constraints.
    # Assignments DF
    data_list = []
    for (sem, jour, am, disc_name, sid) in detailed_assignments:
        data_list.append({
            "Semaine": sem,
            "Jour": jour,
            "Apres-Midi": am, # String usually "True"/"False" or similar depending on CSV
            "Discipline": disc_name,
            "Id_Eleve": sid,
            "Annee": eleves[sid]["annee"], # Add level for Mixite check
            "Jour_Index": ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi"].index(jour) if jour in ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi"] else -1
        })
    
    csv_df = pd.DataFrame(data_list)
    
    if not csv_df.empty:
        # Pre-calculation for "Apres-Midi" boolean if it's string
        # Assuming CSV has "True" or "False" strings or just checking uniqueness

        # 1. Max Vacations / Semaine
        for disc_name, disc in disc_map.items():
            if disc.nb_vacations_par_semaine > 0:
                limit = disc.nb_vacations_par_semaine
                subset = csv_df[csv_df["Discipline"] == disc_name]
                if subset.empty: continue
                
                # Count per student per week
                counts = subset.groupby(["Id_Eleve", "Semaine"]).size().reset_index(name="Count")
                violations = counts[counts["Count"] > limit]
                
                total_checks = len(counts) # checking each student-week instance
                nb_violations = len(violations)
                
                constraint_stats.append({
                    "Discipline": disc_name,
                    "Contrainte": f"Max {limit} / Semaine",
                    "Total_Check": total_checks,
                    "Violations": nb_violations,
                    "Respect": total_checks - nb_violations
                })

        # 2. Fréquence Vacations (Gap check)
        # "Une semaine sur X" -> Gap between weeks >= X
        for disc_name, disc in disc_map.items():
            if disc.frequence_vacations > 1:
                freq = disc.frequence_vacations
                subset = csv_df[csv_df["Discipline"] == disc_name]
                if subset.empty: continue
                
                total_checks = 0
                nb_violations = 0
                
                # Check per student
                for sid, group in subset.groupby("Id_Eleve"):
                    weeks = sorted(group["Semaine"].unique())
                    # Check gaps
                    for i in range(len(weeks) - 1):
                        diff = weeks[i+1] - weeks[i]
                        total_checks += 1
                        # If freq=2 (1 week out of 2), we expect assignment, gap, assignment -> diff >= 2
                        if diff < freq:
                            nb_violations += 1
                            
                constraint_stats.append({
                    "Discipline": disc_name,
                    "Contrainte": f"Fréquence (Gap >= {freq})",
                    "Total_Check": total_checks,
                    "Violations": nb_violations,
                    "Respect": total_checks - nb_violations
                })

        # 3. Répartition Continuité
        # (Limit, Window) e.g. (1, 12) -> Max 1 time in 12 weeks
        for disc_name, disc in disc_map.items():
            if disc.repetition_continuite and isinstance(disc.repetition_continuite, (list, tuple)):
                limit, window = disc.repetition_continuite
                if limit > 0 and window > 0:
                    subset = csv_df[csv_df["Discipline"] == disc_name]
                    if subset.empty: continue

                    total_checks = 0
                    nb_violations = 0
                    
                    # For each student, check rolling window or simple distance
                    # Easier: iterate all weeks and check if violations occur in window
                    # Or check for each assignment, how many others in [w - window + 1, w]
                    
                    for sid, group in subset.groupby("Id_Eleve"):
                        weeks = sorted(group["Semaine"].values)
                        # Sliding window check
                        # For every assigned week, verify no more than 'limit' assignments in 'window' size
                        # Actually the constraint is usually "No more than Limit in ANY Window of size X"
                        
                        # Let's slide a window of size 'window' across the student's timeline
                        if not weeks: continue
                        
                        
                        # We only need to check windows starting at an assigned week (optimization)
                        for start_w in weeks:
                            end_w = start_w + window - 1 
                            # Count assignments in [start_w, end_w]
                            count = sum(1 for w in weeks if start_w <= w <= end_w)
                            
                            total_checks += 1
                            if count > limit:
                                nb_violations += 1
                                # Avoid double counting? A set of bad assignments? 
                                # Let's count "Bad Windows".
                    
                    constraint_stats.append({
                        "Discipline": disc_name,
                        "Contrainte": f"Continuité (Max {limit} en {window} sem)",
                        "Total_Check": total_checks,
                        "Violations": nb_violations,
                        "Respect": total_checks - nb_violations
                    })

        # 4. Mixité des Groupes
        # Check every occupied slot (Week, Day, AM)
        for disc_name, disc in disc_map.items():
            if disc.mixite_groupes > 0:
                mix_type = disc.mixite_groupes
                subset = csv_df[csv_df["Discipline"] == disc_name]
                if subset.empty: continue
                
                # Group by Slot
                grouped = subset.groupby(["Semaine", "Jour", "Apres-Midi"])
                
                total_checks = 0
                nb_violations = 0
                
                for slot, group in grouped:
                    levels = group["Annee"].unique()
                    total_checks += 1
                    
                    if mix_type == 2: # At least 2 levels
                        if len(levels) < 2:
                            nb_violations += 1
                    elif mix_type == 3: # All same level (Exactement 1 level)
                        if len(levels) != 1:
                            nb_violations += 1
                
                constraint_stats.append({
                    "Discipline": disc_name,
                    "Contrainte": f"Mixité (Type {mix_type})",
                    "Total_Check": total_checks,
                    "Violations": nb_violations,
                    "Respect": total_checks - nb_violations
                })

        # 5. Répartition Semestrielle
        # [S1_Quota, S2_Quota] -> S1 = Weeks 1-28, S2 = Weeks 29-52?
        # Model uses: list_periodes = [Periode(0, 34, 44), Periode(1, 45, 51), Periode(2, 1, 8)...]
        # But commonly semesters are 1-26 and 27-52.
        # Let's assume split at 26/27.
        for disc_name, disc in disc_map.items():
            if disc.repartition_semestrielle and len(disc.repartition_semestrielle) == 2:
                q1_req, q2_req = disc.repartition_semestrielle
                subset = csv_df[csv_df["Discipline"] == disc_name]
                if subset.empty: continue
                
                total_checks = 0
                nb_violations = 0
                
                # Check per student
                for sid, group in subset.groupby("Id_Eleve"):
                    # Only check students concerned (those with assignments)
                    # Check S1
                    c1 = group[group["Semaine"] <= 26].shape[0]
                    c2 = group[group["Semaine"] > 26].shape[0]
                    
                    total_checks += 1
                    # Deviation? The constraint usually sets a specific target.
                    # If I have 2 in S1 and 2 in S2.
                    # If assignments < req, is it violation? 
                    # Usually "Repartition" implies balancing.
                    # I will check if c1 matches q1 and c2 matches q2 approximately or strictly?
                    # Since quotas are hard in this model, I'll check exactitude if quota reached.
                    # Or simplistically: |c1 - q1| + |c2 - q2| == 0?
                    # Let's just report violations if they don't meet the target
                    
                    # We only flag if they have enough vacations to meet it?
                    # If a student has total 4, and split is [2,2], then 2-2 is expected.
                    # If they have 3, maybe 2-1 is okay.
                    # Let's assume strict target for now if Total >= Sum(Targets).
                    
                    if c1 != q1_req or c2 != q2_req:
                        # Soften: The constraint is likely: "Try to have X in S1 and Y in S2"
                        nb_violations += 1

                constraint_stats.append({
                    "Discipline": disc_name,
                    "Contrainte": f"Répartition Semestrielle [{q1_req}, {q2_req}]",
                    "Total_Check": total_checks,
                    "Violations": nb_violations,
                    "Respect": total_checks - nb_violations
                })

        # 6. Fill Requirement (Remplissage Obligatoire)
        # Check if slots are filled to capacity if be_filled=True
        
        # Pre-compute counts map for speed
        # Key: (Semaine, Jour, Apres-Midi, Discipline) -> Count
        slot_counts = collections.defaultdict(int)
        for (sem, jour, am, disc_name, sid) in detailed_assignments:
            slot_counts[(sem, jour, am, disc_name)] += 1
            
        jours = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi"]
        
        for disc_name, disc in disc_map.items():
            if disc.be_filled:
                total_checks = 0
                nb_violations = 0
                
                # Iterate all theoretical slots
                for s in range(1, 53):
                    for j_idx, j_str in enumerate(jours):
                        for am in [0, 1]: 
                            am_str = str(am) # CSV stores as string
                            slot_idx = j_idx * 2 + am
                            
                            # Check if slot is open in configuration
                            is_open = (len(disc.presence) > slot_idx and disc.presence[slot_idx])
                            
                            # Expected capacity
                            capacity = disc.nb_eleve[slot_idx] if is_open else 0
                            
                            if capacity > 0:
                                total_checks += 1
                                current_count = slot_counts[(s, j_str, am_str, disc_name)]
                                
                                # Equality check
                                if current_count != capacity:
                                    nb_violations += 1
                
                constraint_stats.append({
                    "Discipline": disc_name,
                    "Contrainte": "Remplissage (Fill Req)",
                    "Total_Check": total_checks,
                    "Violations": nb_violations,
                    "Respect": total_checks - nb_violations
                })

        # 7. Paire Jours (Daily Pairs)
        # Verify that if a student is assigned to day A, they are also assigned to day B in the same week
        for disc_name, disc in disc_map.items():
            if disc.paire_jours:
                subset = csv_df[csv_df["Discipline"] == disc_name]
                if subset.empty: continue
                
                total_checks = 0
                nb_violations = 0
                
                # Check per student/week
                for (sid, week), group in subset.groupby(["Id_Eleve", "Semaine"]):
                    # Get assigned day indices for this week
                    # Map Day String to Index 0-4
                    assigned_days = set(group["Jour_Index"].unique())
                    
                    for (d1, d2) in disc.paire_jours:
                        # Logic: If one exists, the other must exist
                        has_d1 = d1 in assigned_days
                        has_d2 = d2 in assigned_days
                        
                        if has_d1 or has_d2:
                            total_checks += 1
                            if not (has_d1 and has_d2):
                                nb_violations += 1
                
                constraint_stats.append({
                    "Discipline": disc_name,
                    "Contrainte": f"Paire Jours {disc.paire_jours}",
                    "Total_Check": total_checks,
                    "Violations": nb_violations,
                    "Respect": total_checks - nb_violations
                })

        # 8. Respect Jour Préférence
        # Verify assignments match the student's preferred day
        for disc_name, disc in disc_map.items():
            if disc.take_jour_pref:
                subset = csv_df[csv_df["Discipline"] == disc_name]
                if subset.empty: continue
                
                total_checks = 0
                nb_violations = 0
                
                for index, row in subset.iterrows():
                    sid = row["Id_Eleve"]
                    assigned_jour = row["Jour"]
                    
                    # Preference from student data
                    pref_str = eleves[sid].get("jour_preference", "")
                    
                    # If student has a preference (and it's a valid day)
                    if pref_str and pref_str in ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi"]:
                        total_checks += 1
                        if assigned_jour != pref_str:
                            nb_violations += 1
                            
                constraint_stats.append({
                    "Discipline": disc_name,
                    "Contrainte": "Jour Préférence",
                    "Total_Check": total_checks,
                    "Violations": nb_violations,
                    "Respect": total_checks - nb_violations
                })

        # 9. Respect Stages (No assignments during stage weeks)
        if stages:
            total_checks = 0
            violations = 0
            
            # Check per student
            for sid, stud_data in eleves.items():
                annee = stud_data.get("nom_annee")
                p_stage = stud_data.get("periode_stage")
                
                if annee and p_stage and (annee, p_stage) in stages:
                    start, end = stages[(annee, p_stage)]
                    total_checks += 1 # We check this student's schedule
                    
                    # Assignments for this student
                    stud_assignments = csv_df[csv_df["Id_Eleve"] == sid]
                    
                    if not stud_assignments.empty:
                         # Check if any assignment week is in [start, end]
                        conflict = stud_assignments[
                            (stud_assignments["Semaine"] >= start) & 
                            (stud_assignments["Semaine"] <= end)
                        ]
                        
                        if not conflict.empty:
                            violations += 1
            
            constraint_stats.append({
                "Discipline": "TOUTES",
                "Contrainte": "Respect Périodes Stage",
                "Total_Check": total_checks,
                "Violations": violations,
                "Respect": total_checks - violations
            })

    # --- Occupancy Analysis (Remplissage Salles) ---
    occupancy_data = []
    
    if not csv_df.empty:
        # Determine range of weeks
        # Convert Semaine to int just in case
        csv_df["Semaine"] = pd.to_numeric(csv_df["Semaine"], errors='coerce').fillna(0).astype(int)
        all_weeks = sorted(csv_df["Semaine"].unique())
        all_weeks = [w for w in all_weeks if w > 0] # Filter valid weeks
        
        # Pre-compute counts map for speed (Re-use slot_counts if valid, but ensure keys match)
        # slot_counts keys: (Semaine, Jour, Apres-Midi, Discipline)
        # Note: In detailed_assignments, 'Semaine' might be string or int depending on reader. 
        # The reader converts 'id_eleve' but not Semaine explicitly in the loop above?
        # Let's ensure types. In the reader loop:
        # row["Semaine"] is string.
        
        # Let's rebuild a robust counts map from csv_df which is clean
        # csv_df columns: Semaine (int), Jour (str), Apres-Midi (str/int?), Discipline (str)
        
        real_counts = csv_df.groupby(["Semaine", "Jour", "Apres-Midi", "Discipline"]).size().to_dict()
        # Key: (1, 'Lundi', '0', 'Polyclinique')
        
        jours_ord = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi"]
        
        for w in all_weeks:
            for d_idx, day_name in enumerate(jours_ord):
                for am_val in [0, 1]:
                    # Helper index for discipline arrays (0..9)
                    # 0=Lundi Matin, 1=Lundi Soir, 2=Mardi Matin...
                    slot_idx = d_idx * 2 + am_val
                    
                    # AM value in CSV is usually "0" or "1" (String)
                    am_str = str(am_val)
                    
                    for disc_name, disc in disc_map.items():
                        # Check presence (Is the clinic open for this discipline at this time?)
                        if slot_idx < len(disc.presence) and disc.presence[slot_idx]:
                            capacity = disc.nb_eleve[slot_idx]
                            
                            # Get count
                            # Key format matching GroupBy: (Semaine, Jour, Apres-Midi, Discipline)
                            # csv_df["Apres-Midi"] might be int or str. Let's check csv_df dtypes or normalized data.
                            # In loop above: data_list.append(..., "Apres-Midi": am, ...)
                            # am comes from row["Apres-Midi"]. In CSV it is "0" or "1".
                            try:
                                count = real_counts.get((w, day_name, am_str, disc_name), 0)
                            except KeyError:
                                # Try int if key mismatch
                                count = real_counts.get((w, day_name, am_val, disc_name), 0)
                            
                            rate = (count / capacity * 100) if capacity > 0 else 0
                            
                            status = "VIDE"
                            if count == 0: status = "VIDE"
                            elif count == capacity: status = "PLEIN"
                            elif count > capacity: status = "SURCHARGE"
                            else: status = "PARTIEL"
                            
                            occupancy_data.append({
                                "Semaine": w,
                                "Jour": day_name,
                                "Apres-Midi": "Apres-Midi" if am_val == 1 else "Matin",
                                "Discipline": disc_name,
                                "Capacite": capacity,
                                "Occupe": count,
                                "Taux": round(rate, 1),
                                "Status": status
                            })
                            
    df_occupancy = pd.DataFrame(occupancy_data)

    # --- Stats Compilation ---
    stats_data = []
    
    # Per Student/Discipline Stats
    for sid, stud_data in eleves.items():
        annee = stud_data["annee"]
        
        for disc_name, disc_obj in disc_map.items():
            # Check if this student is concerned by this discipline
            if annee not in disc_obj.annee:
                continue
                
            # Get Quota
            try:
                # Assuming simple mapping: index in annee list = index in quota list
                # This depends on logic in model_V4_01.py: adx = disc.annee.index(el.annee.value)
                idx = disc_obj.annee.index(annee)
                quota = disc_obj.quota[idx]
            except ValueError:
                quota = 0
            
            # If quota is 0, is it still tracked?
            if quota == 0:
                # Check if they have assignments anyway
                count = assignments[sid][disc_name]
                if count > 0:
                    stats_data.append({
                        "Id_Eleve": sid,
                        "Annee": stud_data["nom_annee"],
                        "Discipline": disc_name,
                        "Attribue": count,
                        "Objectif": 0,
                        "Pourcentage": "N/A",
                        "Delta": count,
                        "Status": "Hors Quota"
                    })
                continue

            count = assignments[sid][disc_name]
            pct = count / quota * 100
            
            status = "OK"
            if count < quota: status = "Manque"
            elif count > quota: status = "Surplus"

            stats_data.append({
                "Id_Eleve": sid,
                "Annee": stud_data["nom_annee"],
                "Discipline": disc_name,
                "Attribue": count,
                "Objectif": quota,
                "Pourcentage": round(pct, 1),
                "Delta": count - quota,
                "Status": status
            })

    df = pd.DataFrame(stats_data)
    
    # Format constraint stats for export
    df_constraints = pd.DataFrame(constraint_stats)
    if not df_constraints.empty:
        df_constraints["Pourcentage_Respect"] = df_constraints.apply(
            lambda r: round(r["Respect"] / r["Total_Check"] * 100, 1) if r["Total_Check"] > 0 else 100, axis=1
        )

    return df, binome_stats, df_constraints, df_occupancy

def generate_report(df, output_path, binome_stats=None, df_constraints=None, df_occupancy=None, optimization_scores=None):
    wb = Workbook()
    
    # 1. Detailed Sheet
    ws_detail = wb.active
    ws_detail.title = "Détail par Élève"
    for r in dataframe_to_rows(df, index=False, header=True):
        ws_detail.append(r)
    
    # Style logic
    header_style = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    for cell in ws_detail[1]:
        cell.font = header_style
        cell.fill = header_fill
        
    # 2. Summary by Level (Annee) and Discipline
    if not df.empty:
        ws_summary = wb.create_sheet("Synthèse Promo")
        
        # Group by Annee, Discipline -> Sum(Attribue), Sum(Objectif), Mean(Pourcentage)
        
        # Create a cleaning copy for aggregation
        df_agg = df.copy()
        df_agg["Pourcentage"] = pd.to_numeric(df_agg["Pourcentage"], errors='coerce')
        
        summary = df_agg.groupby(["Annee", "Discipline"]).agg({
            "Attribue": "sum",
            "Objectif": "sum",
            "Pourcentage": "mean",
            "Id_Eleve": "count" # Number of students
        }).reset_index()
        
        summary["Moyenne_Vacations"] = summary["Attribue"] / summary["Id_Eleve"]
        summary["Taux_Remplissage_Global"] = (summary["Attribue"] / summary["Objectif"] * 100).round(1)
        summary["Pourcentage"] = summary["Pourcentage"].round(1)
        
        # Rename columns for clarity
        summary = summary.rename(columns={
            "Pourcentage": "Moyenne_%_Individuel",
            "Id_Eleve": "Nb_Eleves"
        })
        
        # Reorder
        cols = ["Annee", "Discipline", "Nb_Eleves", "Attribue", "Objectif", "Moyenne_Vacations", "Moyenne_%_Individuel", "Taux_Remplissage_Global"]
        summary = summary[cols]
        
        for r in dataframe_to_rows(summary, index=False, header=True):
            ws_summary.append(r)

        for cell in ws_summary[1]:
            cell.font = header_style
            cell.fill = header_fill
            
        # 3. Global Discipline Stats
        ws_disc = wb.create_sheet("Synthèse Discipline")
        disc_stats = df.groupby("Discipline").agg({
            "Attribue": "sum",
            "Objectif": "sum"
        }).reset_index()
        disc_stats["Taux_Global"] = (disc_stats["Attribue"] / disc_stats["Objectif"] * 100).round(1)
        
        for r in dataframe_to_rows(disc_stats, index=False, header=True):
            ws_disc.append(r)

        for cell in ws_disc[1]:
            cell.font = header_style
            cell.fill = header_fill

    # 4. General Statistics (Student Totals & Global)
    ws_gen = wb.create_sheet("Synthèse Générale")
    
    # A. Total per Student
    # Group by Id_Eleve, Annee -> Sum(Attribue), Sum(Objectif)
    stud_totals = df.groupby(["Id_Eleve", "Annee"]).agg({
        "Attribue": "sum",
        "Objectif": "sum"
    }).reset_index()
    
    # Calculate global fill rate per student (handle division by zero if objective is 0)
    stud_totals["Taux_Remplissage"] = stud_totals.apply(
        lambda row: round(row["Attribue"] / row["Objectif"] * 100, 1) if row["Objectif"] > 0 else 0, axis=1
    )
    
    ws_gen.append(["Statistiques par Élève (Toutes disciplines confondues)"])
    ws_gen.append(["Id_Eleve", "Annee", "Total_Vacations", "Total_Objectif", "Taux_Remplissage"])
    
    # Style header
    for col in range(1, 6):
        cell = ws_gen.cell(row=2, column=col)
        cell.font = header_style
        cell.fill = header_fill

    for r in dataframe_to_rows(stud_totals, index=False, header=False):
        ws_gen.append(r)
        
    ws_gen.append([]) # Spacer
    ws_gen.append([]) # Spacer

    # B. Total per Level (Promo)
    # Group by Annee -> Sum(Attribue), Sum(Objectif), Count(Unique Students)
    level_totals = df.groupby("Annee").agg({
        "Attribue": "sum",
        "Objectif": "sum",
        "Id_Eleve": "nunique"
    }).reset_index()
    
    level_totals["Moyenne_Vacations"] = (level_totals["Attribue"] / level_totals["Id_Eleve"]).round(1)
    level_totals["Taux_Global"] = (level_totals["Attribue"] / level_totals["Objectif"] * 100).round(1)

    start_row = ws_gen.max_row + 1
    ws_gen.append(["Statistiques par Promotion"])
    ws_gen.append(["Annee", "Total_Vacations_Promo", "Total_Objectif_Promo", "Nb_Eleves", "Moyenne_Par_Eleve", "Taux_Global"])
    
    # Style header
    for col in range(1, 7):
        cell = ws_gen.cell(row=start_row+1, column=col)
        cell.font = header_style
        cell.fill = header_fill

    for r in dataframe_to_rows(level_totals, index=False, header=False):
        ws_gen.append(r)

    ws_gen.append([]) # Spacer
    ws_gen.append([]) # Spacer

    # C. Grand Total
    grand_total_attrib = df["Attribue"].sum()
    grand_total_obj = df["Objectif"].sum()
    grand_total_pct = round(grand_total_attrib / grand_total_obj * 100, 1) if grand_total_obj > 0 else 0
    
    start_row = ws_gen.max_row + 1
    ws_gen.append(["Statistiques Globales (Tout le planning)"])
    ws_gen.append(["Total_Vacations_Affectées", "Total_Quota_Attendu", "Taux_Global"])
    
    # Style header
    for col in range(1, 4):
        cell = ws_gen.cell(row=start_row+1, column=col)
        cell.font = header_style
        cell.fill = header_fill

    ws_gen.append([grand_total_attrib, grand_total_obj, grand_total_pct])

    # Adjust column widths
    ws_gen.column_dimensions['A'].width = 25
    ws_gen.column_dimensions['B'].width = 25
    ws_gen.column_dimensions['C'].width = 25
    ws_gen.column_dimensions['D'].width = 25
    ws_gen.column_dimensions['E'].width = 25

    # 5. Binome Stats
    if binome_stats:
        ws_binome = wb.create_sheet("Stats Binômes")
        ws_binome.append(["Statistiques Respect des Binômes"])
        ws_binome.append(["Métrique", "Valeur"])
        
        # Style
        ws_binome.cell(row=2, column=1).font = header_style
        ws_binome.cell(row=2, column=1).fill = header_fill
        ws_binome.cell(row=2, column=2).font = header_style
        ws_binome.cell(row=2, column=2).fill = header_fill
        
        ws_binome.append(["Total Vacations (Disciplines en Binome)", binome_stats["Total_Binome_Vacations_Checked"]])
        ws_binome.append(["Dont Binomes Respectés (Paires)", binome_stats["Respect_Binome"]])
        ws_binome.append(["Dont Binomes Violés (Seuls)", binome_stats["Violations"]])
        ws_binome.append(["Taux de Respect (%)", f"{binome_stats['Percentage']}%"])
        
        ws_binome.column_dimensions['A'].width = 40
        ws_binome.column_dimensions['B'].width = 15

    # 6. Constraints Details
    if df_constraints is not None and not df_constraints.empty:
        ws_constr = wb.create_sheet("Stats Contraintes")
        
        # Header
        cols = ["Discipline", "Contrainte", "Total_Check", "Violations", "Respect", "Pourcentage_Respect"]
        ws_constr.append(cols)
        
        for cell in ws_constr[1]:
            cell.font = header_style
            cell.fill = header_fill

        # Data
        for r in dataframe_to_rows(df_constraints[cols], index=False, header=False):
            ws_constr.append(r)
            
        # Formatting
        ws_constr.column_dimensions['A'].width = 20
        ws_constr.column_dimensions['B'].width = 35
        ws_constr.column_dimensions['C'].width = 12
        ws_constr.column_dimensions['D'].width = 12
        ws_constr.column_dimensions['E'].width = 12
        ws_constr.column_dimensions['F'].width = 15
        
        # Color coding for %
        # (Advanced styling omitted for brevity, simple list is fine)

    # 7. Occupancy Stats (Remplissage Salles)
    if df_occupancy is not None and not df_occupancy.empty:
        ws_occ = wb.create_sheet("Remplissage Salles")
        
        # A. Summary by Discipline
        ws_occ.append(["Synthèse Remplissage par Discipline"])
        
        # Group: Discipline -> Sum(Occupe), Sum(Capacite), Avg(Taux), Count(Status=PLEIN)...
        summary_occ = df_occupancy.groupby("Discipline").agg({
            "Occupe": "sum",
            "Capacite": "sum",
            "Taux": "mean",
            "Semaine": "count" # Total slots open
        }).reset_index()
        
        summary_occ["Taux_Global"] = (summary_occ["Occupe"] / summary_occ["Capacite"] * 100).round(1)
        summary_occ = summary_occ.rename(columns={"Semaine": "Nb_Creneaux_Ouverts", "Taux": "Moyenne_Taux_Creneau"})
        summary_occ["Moyenne_Taux_Creneau"] = summary_occ["Moyenne_Taux_Creneau"].round(1)
        
        # Calculate Status Counts
        status_counts = df_occupancy.groupby(["Discipline", "Status"]).size().unstack(fill_value=0).reset_index()
        # Ensure cols exist
        for col in ["VIDE", "PARTIEL", "PLEIN", "SURCHARGE"]:
            if col not in status_counts.columns:
                status_counts[col] = 0
                
        # Merge
        summary_occ = pd.merge(summary_occ, status_counts, on="Discipline", how="left")
        
        cols = ["Discipline", "Nb_Creneaux_Ouverts", "Capacite", "Occupe", "Taux_Global", "Moyenne_Taux_Creneau", "VIDE", "PARTIEL", "PLEIN", "SURCHARGE"]
        ws_occ.append(cols)
        
        for cell in ws_occ[2]:
            cell.font = header_style
            cell.fill = header_fill

        for r in dataframe_to_rows(summary_occ[cols], index=False, header=False):
            ws_occ.append(r)
            
        ws_occ.append([])
        ws_occ.append([])
        
        # B. Detailed List
        ws_occ.append(["Détail par Créneau"])
        detail_header_row = ws_occ.max_row
        ws_occ.append(list(df_occupancy.columns))
        
        for cell in ws_occ[detail_header_row + 1]:
            cell.font = header_style
            cell.fill = header_fill
            
        for r in dataframe_to_rows(df_occupancy, index=False, header=False):
            ws_occ.append(r)
            
        ws_occ.column_dimensions['A'].width = 10
        ws_occ.column_dimensions['B'].width = 15
        ws_occ.column_dimensions['C'].width = 12
        ws_occ.column_dimensions['D'].width = 20
    # 8. Optimization Scores (if available)
    if optimization_scores:
        ws_scores = wb.create_sheet("Scores Optimisation")
        ws_scores.append(["Métrique", "Valeur"])
        
        # Style header
        for col in range(1, 3):
            cell = ws_scores.cell(1, col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4F81BD")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add scores data
        ws_scores.append(["Score Brut", f"{optimization_scores.get('raw_score', 0):,.0f}"])
        ws_scores.append(["Score Maximum Théorique", f"{optimization_scores.get('max_theoretical_score', 0):,.0f}"])
        ws_scores.append(["Score Normalisé (/100)", f"{optimization_scores.get('normalized_score', 0):.2f}"])
        ws_scores.append(["Statut Solution", optimization_scores.get('status', 'UNKNOWN')])
        
        # Format
        ws_scores.column_dimensions['A'].width = 30
        ws_scores.column_dimensions['B'].width = 25
        
        # Highlight normalized score
        score_cell = ws_scores.cell(4, 2)  # Row 4 is normalized score
        score_val = optimization_scores.get('normalized_score', 0)
        if score_val >= 80:
            score_cell.fill = PatternFill("solid", fgColor="90EE90")  # Light green
        elif score_val >= 60:
            score_cell.fill = PatternFill("solid", fgColor="FFFF99")  # Light yellow
        else:
            score_cell.fill = PatternFill("solid", fgColor="FFB6C1")  # Light red
    # Save
    wb.save(output_path)
    print(f"Report generated: {output_path}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", "-i", type=str, help="Input CSV solution file")
    parser.add_argument("--output", "-o", type=str, help="Output Excel statistics file")
    args = parser.parse_args()

    base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
    
    solution_file = args.input if args.input else os.path.join(base_dir, 'resultat', 'planning_solution.csv')
    output_file = args.output if args.output else os.path.join(base_dir, 'resultat', 'statistiques_solution.xlsx')
    
    # Load optimization scores if available
    # Try multiple locations: same folder as CSV, or default resultat folder
    optimization_scores = None
    scores_locations = []
    
    if args.input:
        # Si input fourni, chercher le JSON dans le même dossier
        input_dir = os.path.dirname(os.path.abspath(args.input))
        input_basename = os.path.basename(args.input).replace('.csv', '')
        scores_locations.append(os.path.join(input_dir, 'optimization_scores.json'))
        scores_locations.append(os.path.join(input_dir, f'{input_basename}_scores.json'))
    
    # Toujours ajouter le dossier par défaut
    scores_locations.append(os.path.join(base_dir, 'resultat', 'optimization_scores.json'))
    
    # Essayer chaque emplacement
    for scores_file in scores_locations:
        if os.path.exists(scores_file):
            try:
                with open(scores_file, 'r', encoding='utf-8') as f:
                    optimization_scores = json.load(f)
                print(f"Scores d'optimisation chargés depuis {scores_file}")
                break
            except Exception as e:
                print(f"Erreur lors du chargement de {scores_file} : {e}")
    
    disc_map, eleves, stages = load_data()
    if disc_map and eleves:
        print("Analyzing solution...")
        df, binome_stats, df_constraints, df_occupancy = analyze_solution(solution_file, disc_map, eleves, stages)
        if df is not None:
            generate_report(df, output_file, binome_stats, df_constraints, df_occupancy, optimization_scores)
            
            # Sauvegarder les scores dans un fichier JSON à côté de l'Excel
            if optimization_scores:
                json_output = output_file.replace('.xlsx', '_scores.json')
                try:
                    with open(json_output, 'w', encoding='utf-8') as f:
                        json.dump(optimization_scores, f, indent=2)
                    print(f"Scores sauvegardés dans: {json_output}")
                except Exception as e:
                    print(f"Erreur lors de la sauvegarde des scores JSON : {e}")
            
            print("Done.")
