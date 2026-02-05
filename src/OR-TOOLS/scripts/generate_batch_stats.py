import os
import sys
import subprocess
import re
import json
from pathlib import Path

# Add parent directories to sys.path
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent.parent
sys.path.append(str(PROJECT_ROOT))

STATS_SCRIPT = PROJECT_ROOT / "src" / "analysis" / "generate_statistics.py"

def parse_scores_from_log(log_file):
    """
    Extrait les scores d'optimisation depuis un fichier de log.
    
    Cherche des lignes comme:
    Score brut : 155,290,666,940
    Score maximum théorique : 290,808,782,540
    Score normalisé : 53.40/100
    
    Returns:
        dict ou None si les scores ne sont pas trouvés
    """
    if not log_file.exists():
        return None
    
    try:
        with open(log_file, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # Patterns pour extraire les scores
        # Score brut peut avoir des virgules comme séparateurs de milliers
        raw_pattern = r"Score brut\s*:\s*([\d,]+)"
        max_pattern = r"Score maximum théorique\s*:\s*([\d,]+)"
        norm_pattern = r"Score normalisé\s*:\s*([\d.]+)/100"
        
        raw_match = re.search(raw_pattern, content)
        max_match = re.search(max_pattern, content)
        norm_match = re.search(norm_pattern, content)
        
        if raw_match and max_match and norm_match:
            # Nettoyer les virgules des nombres
            raw_score = int(raw_match.group(1).replace(',', ''))
            max_score = int(max_match.group(1).replace(',', ''))
            norm_score = float(norm_match.group(1))
            
            # Détecter le statut (OPTIMAL ou FEASIBLE)
            # Chercher "status: OPTIMAL" ou "status: FEASIBLE" (format exact du log)
            status = "FEASIBLE"
            status_pattern = r"status:\s*(OPTIMAL|FEASIBLE)"
            status_match = re.search(status_pattern, content, re.IGNORECASE)
            
            if status_match:
                status = status_match.group(1).upper()
            
            return {
                "raw_score": raw_score,
                "max_theoretical_score": max_score,
                "normalized_score": norm_score,
                "status": status
            }
        else:
            return None
            
    except Exception as e:
        print(f"  ⚠ Erreur lors du parsing du log : {e}")
        return None

def generate_stats_for_iterations(base_folder):
    """
    Parcourt tous les dossiers d'itérations et génère les statistiques
    pour chaque fichier CSV trouvé.
    
    Args:
        base_folder: Dossier de base contenant les sous-dossiers (ex: T1200/V5_01/)
    """
    base_path = Path(base_folder)
    
    if not base_path.exists():
        print(f"Erreur: Le dossier {base_folder} n'existe pas.")
        return
    
    # Chercher le dossier iters/
    iters_folder = base_path / "iters"
    stats_folder = base_path / "stats"
    logs_folder = base_path / "logs"
    
    if not iters_folder.exists():
        print(f"Erreur: Le dossier {iters_folder} n'existe pas.")
        return
    
    # Créer le dossier stats s'il n'existe pas
    stats_folder.mkdir(parents=True, exist_ok=True)
    
    # Trouver tous les fichiers CSV
    csv_files = sorted(iters_folder.glob("*.csv"))
    
    if not csv_files:
        print(f"Aucun fichier CSV trouvé dans {iters_folder}")
        return
    
    print(f"Trouvé {len(csv_files)} fichiers CSV dans {iters_folder}")
    print(f"Génération des statistiques dans {stats_folder}")
    print("=" * 70)
    
    success_count = 0
    failed_count = 0
    
    for csv_file in csv_files:
        # Générer le nom du fichier de sortie
        base_name = csv_file.stem  # nom sans extension
        stats_file = stats_folder / f"stats_{base_name}.xlsx"
        
        # Chercher le fichier log correspondant
        log_file = logs_folder / f"log_{base_name}.txt"
        
        print(f"\n[{success_count + failed_count + 1}/{len(csv_files)}] {csv_file.name}")
        print(f"  → CSV: {csv_file}")
        print(f"  → Stats: {stats_file.name}")
        
        # Parser les scores depuis le log
        optimization_scores = None
        if log_file.exists():
            print(f"  → Log: {log_file.name}")
            optimization_scores = parse_scores_from_log(log_file)
            if optimization_scores:
                print(f"  ✓ Scores extraits du log: {optimization_scores['normalized_score']:.2f}/100")
            else:
                print(f"  ⚠ Impossible d'extraire les scores du log")
        else:
            print(f"  ⚠ Pas de fichier log trouvé")
        
        # Créer un fichier JSON temporaire avec les scores si disponibles
        temp_json = None
        if optimization_scores:
            temp_json = csv_file.parent / f"{base_name}_scores.json"
            try:
                with open(temp_json, 'w', encoding='utf-8') as f:
                    json.dump(optimization_scores, f, indent=2)
            except Exception as e:
                print(f"  ⚠ Erreur lors de la création du JSON temporaire : {e}")
                temp_json = None
        
        try:
            # Construire la commande
            cmd = [
                sys.executable,
                str(STATS_SCRIPT),
                "--input", str(csv_file),
                "--output", str(stats_file)
            ]
            
            # Exécuter le script de statistiques
            result = subprocess.run(
                cmd,
                check=True,
                capture_output=True,
                text=True,
                cwd=str(PROJECT_ROOT)
            )
            
            print(f"  ✓ Statistiques générées avec succès")
            success_count += 1
            
            # Nettoyer le fichier JSON temporaire
            if temp_json and temp_json.exists():
                try:
                    temp_json.unlink()
                except:
                    pass
            
        except subprocess.CalledProcessError as e:
            print(f"  ✗ ERREUR lors de la génération des statistiques")
            print(f"    {e.stderr if e.stderr else e.stdout}")
            failed_count += 1
            
            # Nettoyer le fichier JSON temporaire en cas d'erreur aussi
            if temp_json and temp_json.exists():
                try:
                    temp_json.unlink()
                except:
                    pass
                    
        except Exception as e:
            print(f"  ✗ ERREUR: {str(e)}")
            failed_count += 1
            
            # Nettoyer le fichier JSON temporaire en cas d'erreur aussi
            if temp_json and temp_json.exists():
                try:
                    temp_json.unlink()
                except:
                    pass
    
    print("\n" + "=" * 70)
    print(f"Terminé: {success_count} réussis, {failed_count} échoués")
    print(f"Statistiques sauvegardées dans: {stats_folder}")
    print("=" * 70)
    
    # Générer la synthèse des scores
    generate_scores_summary(stats_folder, base_path)

def generate_scores_summary(stats_folder, base_path):
    """
    Génère un fichier de synthèse avec les statistiques des scores (moyenne, min, max, écart-type).
    Collecte les scores depuis les logs et génère des fichiers JSON et Excel de synthèse.
    
    Args:
        stats_folder: Dossier contenant les fichiers de stats
        base_path: Dossier de base pour déterminer le nom de configuration
    """
    import statistics
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, LineChart, Reference
    
    # Trouver le dossier logs
    logs_folder = base_path / "logs"
    
    if not logs_folder.exists():
        print("\n  ⚠ Dossier logs introuvable pour la synthèse")
        return
    
    # Collecter les scores depuis les logs
    scores_data = {
        "raw_scores": [],
        "max_theoretical_scores": [],
        "normalized_scores": [],
        "statuses": [],
        "iterations": []
    }
    
    log_files = sorted(logs_folder.glob("log_*.txt"))
    
    if not log_files:
        print("\n  ⚠ Aucun fichier log trouvé pour la synthèse")
        return
    
    print(f"\n  📊 Collecte des scores depuis {len(log_files)} fichiers logs...")
    
    for idx, log_file in enumerate(log_files, 1):
        try:
            optimization_scores = parse_scores_from_log(log_file)
            if optimization_scores:
                scores_data["raw_scores"].append(optimization_scores["raw_score"])
                scores_data["max_theoretical_scores"].append(optimization_scores["max_theoretical_score"])
                scores_data["normalized_scores"].append(optimization_scores["normalized_score"])
                scores_data["statuses"].append(optimization_scores["status"])
                scores_data["iterations"].append(idx)
        except Exception as e:
            print(f"  ⚠ Erreur lors de la lecture de {log_file.name}: {e}")
            continue
    
    if not scores_data["raw_scores"]:
        print("\n  ⚠ Aucun score valide trouvé")
        return
    
    # Calculer les statistiques
    n = len(scores_data["raw_scores"])
    
    summary = {
        "configuration": {
            "model": base_path.name,  # V5_01, V5_02, etc.
            "time_limit": base_path.parent.name,  # T1200, T1800, etc.
            "date": base_path.parent.parent.name,  # Date folder
            "iterations_count": n
        },
        "raw_score": {
            "mean": round(statistics.mean(scores_data["raw_scores"]), 2),
            "median": round(statistics.median(scores_data["raw_scores"]), 2),
            "min": min(scores_data["raw_scores"]),
            "max": max(scores_data["raw_scores"]),
            "stdev": round(statistics.stdev(scores_data["raw_scores"]), 2) if n > 1 else 0
        },
        "max_theoretical_score": {
            "mean": round(statistics.mean(scores_data["max_theoretical_scores"]), 2),
            "median": round(statistics.median(scores_data["max_theoretical_scores"]), 2),
            "min": min(scores_data["max_theoretical_scores"]),
            "max": max(scores_data["max_theoretical_scores"])
        },
        "normalized_score": {
            "mean": round(statistics.mean(scores_data["normalized_scores"]), 2),
            "median": round(statistics.median(scores_data["normalized_scores"]), 2),
            "min": round(min(scores_data["normalized_scores"]), 2),
            "max": round(max(scores_data["normalized_scores"]), 2),
            "stdev": round(statistics.stdev(scores_data["normalized_scores"]), 2) if n > 1 else 0
        },
        "status": {
            "OPTIMAL": scores_data["statuses"].count("OPTIMAL"),
            "FEASIBLE": scores_data["statuses"].count("FEASIBLE")
        },
        "detailed_iterations": [
            {
                "iteration": i,
                "raw_score": r,
                "max_theoretical": m,
                "normalized_score": n,
                "status": s
            }
            for i, r, m, n, s in zip(
                scores_data["iterations"],
                scores_data["raw_scores"],
                scores_data["max_theoretical_scores"],
                scores_data["normalized_scores"],
                scores_data["statuses"]
            )
        ]
    }
    
    # Sauvegarder le fichier JSON de synthèse
    summary_file = base_path / "scores_summary.json"
    try:
        with open(summary_file, 'w', encoding='utf-8') as f:
            json.dump(summary, f, indent=2, ensure_ascii=False)
        print(f"\n  ✓ Synthèse JSON sauvegardée: {summary_file.name}")
    except Exception as e:
        print(f"\n  ✗ Erreur lors de la sauvegarde de la synthèse JSON: {e}")
        return
    
    # Générer un fichier Excel de synthèse
    excel_summary_file = base_path / "scores_summary.xlsx"
    try:
        wb = Workbook()
        
        # === FEUILLE 1: Vue d'ensemble ===
        ws_overview = wb.active
        ws_overview.title = "Vue d'ensemble"
        
        # Style headers
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="366092")
        title_font = Font(bold=True, size=14, color="1F4E78")
        
        # Configuration
        ws_overview.append(["SYNTHÈSE DES SCORES D'OPTIMISATION"])
        ws_overview["A1"].font = Font(bold=True, size=16, color="1F4E78")
        ws_overview.append([])
        
        ws_overview.append(["Configuration"])
        ws_overview["A3"].font = title_font
        ws_overview.append(["Modèle", summary["configuration"]["model"]])
        ws_overview.append(["Limite de temps", summary["configuration"]["time_limit"]])
        ws_overview.append(["Date", summary["configuration"]["date"]])
        ws_overview.append(["Nombre d'itérations", summary["configuration"]["iterations_count"]])
        ws_overview.append([])
        
        # Statistiques Score Normalisé
        ws_overview.append(["Score Normalisé (/100)"])
        ws_overview["A9"].font = title_font
        ws_overview.append(["Métrique", "Valeur"])
        ws_overview["A10"].font = header_font
        ws_overview["A10"].fill = header_fill
        ws_overview["B10"].font = header_font
        ws_overview["B10"].fill = header_fill
        
        ws_overview.append(["Moyenne", summary["normalized_score"]["mean"]])
        ws_overview.append(["Médiane", summary["normalized_score"]["median"]])
        ws_overview.append(["Minimum", summary["normalized_score"]["min"]])
        ws_overview.append(["Maximum", summary["normalized_score"]["max"]])
        ws_overview.append(["Écart-type", summary["normalized_score"]["stdev"]])
        ws_overview.append([])
        
        # Statistiques Score Brut
        ws_overview.append(["Score Brut"])
        ws_overview["A17"].font = title_font
        ws_overview.append(["Métrique", "Valeur"])
        ws_overview["A18"].font = header_font
        ws_overview["A18"].fill = header_fill
        ws_overview["B18"].font = header_font
        ws_overview["B18"].fill = header_fill
        
        ws_overview.append(["Moyenne", f"{summary['raw_score']['mean']:,.2f}"])
        ws_overview.append(["Médiane", f"{summary['raw_score']['median']:,.2f}"])
        ws_overview.append(["Minimum", f"{summary['raw_score']['min']:,}"])
        ws_overview.append(["Maximum", f"{summary['raw_score']['max']:,}"])
        ws_overview.append(["Écart-type", f"{summary['raw_score']['stdev']:,.2f}"])
        ws_overview.append([])
        
        # Statut des solutions
        ws_overview.append(["Statut des Solutions"])
        ws_overview["A25"].font = title_font
        ws_overview.append(["Statut", "Nombre"])
        ws_overview["A26"].font = header_font
        ws_overview["A26"].fill = header_fill
        ws_overview["B26"].font = header_font
        ws_overview["B26"].fill = header_fill
        
        ws_overview.append(["OPTIMAL", summary["status"]["OPTIMAL"]])
        ws_overview.append(["FEASIBLE", summary["status"]["FEASIBLE"]])
        
        # Colorer les cellules selon le score moyen
        mean_score = summary["normalized_score"]["mean"]
        mean_cell = ws_overview["B11"]
        if mean_score >= 80:
            mean_cell.fill = PatternFill("solid", fgColor="90EE90")  # Vert
        elif mean_score >= 60:
            mean_cell.fill = PatternFill("solid", fgColor="FFFF99")  # Jaune
        else:
            mean_cell.fill = PatternFill("solid", fgColor="FFB6C1")  # Rouge
        
        # Largeur des colonnes
        ws_overview.column_dimensions['A'].width = 30
        ws_overview.column_dimensions['B'].width = 20
        
        # === FEUILLE 2: Détail des itérations ===
        ws_detail = wb.create_sheet("Détail Itérations")
        
        ws_detail.append(["Itération", "Score Brut", "Score Max Théorique", "Score Normalisé", "Statut"])
        for cell in ws_detail[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        for iteration in summary["detailed_iterations"]:
            ws_detail.append([
                iteration["iteration"],
                iteration["raw_score"],
                iteration["max_theoretical"],
                iteration["normalized_score"],
                iteration["status"]
            ])
            
            # Colorer la cellule du score normalisé
            row_idx = ws_detail.max_row
            score_cell = ws_detail.cell(row_idx, 4)
            score_val = iteration["normalized_score"]
            
            if score_val >= 80:
                score_cell.fill = PatternFill("solid", fgColor="90EE90")
            elif score_val >= 60:
                score_cell.fill = PatternFill("solid", fgColor="FFFF99")
            else:
                score_cell.fill = PatternFill("solid", fgColor="FFB6C1")
        
        # Largeur des colonnes
        ws_detail.column_dimensions['A'].width = 12
        ws_detail.column_dimensions['B'].width = 18
        ws_detail.column_dimensions['C'].width = 20
        ws_detail.column_dimensions['D'].width = 18
        ws_detail.column_dimensions['E'].width = 12
        
        # === FEUILLE 3: Graphiques ===
        ws_charts = wb.create_sheet("Graphiques")
        
        # Graphique en barres des scores normalisés par itération
        chart1 = BarChart()
        chart1.title = "Score Normalisé par Itération"
        chart1.y_axis.title = "Score (/100)"
        chart1.x_axis.title = "Itération"
        
        # Références aux données
        data = Reference(ws_detail, min_col=4, min_row=1, max_row=len(summary["detailed_iterations"]) + 1)
        cats = Reference(ws_detail, min_col=1, min_row=2, max_row=len(summary["detailed_iterations"]) + 1)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        chart1.height = 12
        chart1.width = 24
        
        ws_charts.add_chart(chart1, "A2")
        
        # Graphique en ligne de l'évolution
        chart2 = LineChart()
        chart2.title = "Évolution du Score Normalisé"
        chart2.y_axis.title = "Score (/100)"
        chart2.x_axis.title = "Itération"
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
        chart2.height = 12
        chart2.width = 24
        
        ws_charts.add_chart(chart2, "A28")
        
        wb.save(excel_summary_file)
        print(f"  ✓ Synthèse Excel sauvegardée: {excel_summary_file.name}")
        
    except Exception as e:
        print(f"\n  ✗ Erreur lors de la création du fichier Excel: {e}")
    
    # Afficher un résumé dans le terminal
    print(f"\n  📊 RÉSUMÉ DES SCORES ({n} itérations):")
    print(f"     ═══════════════════════════════════════")
    print(f"     Score normalisé moyen    : {summary['normalized_score']['mean']:.2f}/100")
    print(f"     Score normalisé médian   : {summary['normalized_score']['median']:.2f}/100")
    print(f"     Plage                    : [{summary['normalized_score']['min']:.2f}, {summary['normalized_score']['max']:.2f}]")
    print(f"     Écart-type              : {summary['normalized_score']['stdev']:.2f}")
    print(f"     ───────────────────────────────────────")
    print(f"     Solutions OPTIMAL       : {summary['status']['OPTIMAL']}")
    print(f"     Solutions FEASIBLE      : {summary['status']['FEASIBLE']}")
    print(f"     ═══════════════════════════════════════")

def main():
    import argparse
    
    parser = argparse.ArgumentParser(
        description="Génère les statistiques pour tous les fichiers CSV d'itérations"
    )
    parser.add_argument(
        "folder",
        type=str,
        help="Chemin vers le dossier contenant les sous-dossiers iters/ et stats/ (ex: batch_experiments/04_02_2026/T1200/V5_01)"
    )
    
    args = parser.parse_args()
    
    generate_stats_for_iterations(args.folder)

if __name__ == "__main__":
    main()
