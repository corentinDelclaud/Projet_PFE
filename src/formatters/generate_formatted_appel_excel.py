import csv
import os
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_header_style():
    return {
        'font': Font(bold=True, color="FFFFFF"),
        'fill': PatternFill(start_color="808080", end_color="808080", fill_type="solid"),
        'alignment': Alignment(horizontal="center", vertical="center"),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    }

def create_cell_style():
    return {
        'alignment': Alignment(vertical="center"),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    }

def generate_appel_excel(input_path: str, output_path: str, target_semaine: int = None, target_jour: str = None, target_periode: str = None, target_discipline: str = None):
    """
    Generates an Excel attendance sheet from the planning CSV file.
    Groups entries by Discipline, Day, and Period.
    Can filter by specific Week, Day, Period, or Discipline.
    """
    print(f"Generating Excel from: {input_path}")
    
    if not os.path.exists(input_path):
        print(f"Error: Input file {input_path} not found.")
        return

    # 1. Read CSV Data and Organize by Session
    sessions = {}
    
    try:
        with open(input_path, mode='r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                if "STAGE:" in row["Discipline"]:
                    continue

                semaine = int(row["Semaine"])
                jour = row["Jour"]
                periode = row["Periode"]
                discipline = row["Discipline"]
                eleve = row["Eleve"]

                if semaine not in sessions:
                    sessions[semaine] = {}
                if jour not in sessions[semaine]:
                    sessions[semaine][jour] = {}
                if periode not in sessions[semaine][jour]:
                    sessions[semaine][jour][periode] = {}
                if discipline not in sessions[semaine][jour][periode]:
                    sessions[semaine][jour][periode][discipline] = []
                
                sessions[semaine][jour][periode][discipline].append(eleve)
    except Exception as e:
        print(f"Error reading CSV: {e}")
        return

    # 2. Build Excel Document
    wb = Workbook()
    ws = wb.active
    ws.title = "Feuilles d'Appel"
    
    # Set Column Widths
    ws.column_dimensions['A'].width = 30  # Nom Étudiant
    ws.column_dimensions['B'].width = 15  # Présent
    ws.column_dimensions['C'].width = 30  # Signature
    ws.column_dimensions['D'].width = 30  # Observations

    header_style = create_header_style()
    cell_style = create_cell_style()

    row_idx = 1

    sorted_semaines = sorted(sessions.keys())
    day_order = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi"]
    period_order = ["Matin", "Apres-Midi"]

    found_any = False

    for semaine in sorted_semaines:
        if semaine not in sessions: continue
        if target_semaine is not None and semaine != target_semaine: continue
        
        # Week Header (Optional, looks nice to separate)
        ws.cell(row=row_idx, column=1, value=f"SEMAINE {semaine}").font = Font(size=14, bold=True)
        row_idx += 2

        for jour in day_order:
            if jour not in sessions[semaine]: continue
            if target_jour is not None and jour != target_jour: continue
            
            for periode in period_order:
                if periode not in sessions[semaine][jour]: continue
                if target_periode is not None and periode != target_periode: continue
                
                disciplines_dict = sessions[semaine][jour][periode]
                if not disciplines_dict: continue

                active_disciplines = [d for d in disciplines_dict if target_discipline is None or d == target_discipline]
                if not active_disciplines: continue

                found_any = True
                
                # Section Header
                ws.cell(row=row_idx, column=1, value=f"{jour} - {periode}").font = Font(size=12, bold=True)
                row_idx += 2

                for discipline in active_disciplines:
                    students = disciplines_dict[discipline]
                    
                    # Discipline Header
                    ws.cell(row=row_idx, column=1, value=discipline).font = Font(bold=True)
                    row_idx += 1
                    
                    # Table Headers
                    headers = ['Nom Étudiant', 'Présent', 'Signature', 'Observations']
                    for col_num, header in enumerate(headers, 1):
                        cell = ws.cell(row=row_idx, column=col_num, value=header)
                        cell.font = header_style['font']
                        cell.fill = header_style['fill']
                        cell.alignment = header_style['alignment']
                        cell.border = header_style['border']
                    row_idx += 1
                    
                    # Student Rows
                    for student in sorted(students):
                        ws.cell(row=row_idx, column=1, value=student).border = cell_style['border']
                        ws.cell(row=row_idx, column=2).border = cell_style['border']
                        ws.cell(row=row_idx, column=3).border = cell_style['border']
                        ws.cell(row=row_idx, column=4).border = cell_style['border']
                        row_idx += 1
                    
                    row_idx += 2 # Space between tables

    if not found_any:
        print("Warning: No sessions found matching criteria.")
        return

    try:
        wb.save(output_path)
        print(f"Excel successfully generated at: {output_path}")
    except Exception as e:
        print(f"Error saving Excel: {e}")

def generate_all_discipline_appel_excels_for_one_day(input_dir: str, output_dir: str, target_semaine: int, target_jour: str, target_periode: str = None):
    """
    Scans input_dir for CSV files matching the criteria (Week/Day) and generates one Excel file per CSV.
    Assumes filename format: Semaine{S}_{Jour}_{Periode}_{Discipline}.csv
    """
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"Created output directory: {output_dir}")

    print(f"Scanning {input_dir} for Semaine {target_semaine}, Jour {target_jour}...")
    
    if not os.path.exists(input_dir):
        print(f"Error: Input directory {input_dir} does not exist.")
        return

    found_any = False
    header_style = create_header_style()
    cell_style = create_cell_style()

    for filename in os.listdir(input_dir):
        if not filename.endswith(".csv"): continue
        
        parts = filename.replace(".csv", "").split("_")
        if len(parts) < 4: continue 
        
        file_semaine_str = parts[0]
        file_jour = parts[1]
        file_periode = parts[2]
        file_discipline = "_".join(parts[3:])
        
        try:
            file_semaine = int(file_semaine_str.replace("Semaine", ""))
        except:
            continue
            
        if file_semaine != target_semaine: continue
        if file_jour != target_jour: continue
        if target_periode and file_periode != target_periode: continue
        
        found_any = True
        
        students = []
        csv_full_path = os.path.join(input_dir, filename)
        try:
            with open(csv_full_path, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    if "Eleve" in row and row["Eleve"]:
                        students.append(row["Eleve"])
        except Exception as e:
            print(f"Failed to read {filename}: {e}")
            continue

        if not students:
            continue

        # Generate Excel
        output_excel_name = f"feuille_appel_{filename.replace('.csv', '.xlsx')}"
        output_excel_path = os.path.join(output_dir, output_excel_name)
        
        print(f"Generating Excel for {file_discipline}...")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Appel" 
        
        # Setup Page Layout for printing
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 30

        row_idx = 1
        
        # Title
        title_cell = ws.cell(row=row_idx, column=1, value=f"Feuille d'Appel - {file_discipline.replace('_', ' ')}")
        title_cell.font = Font(size=14, bold=True)
        row_idx += 1
        
        subtitle_cell = ws.cell(row=row_idx, column=1, value=f"Semaine {file_semaine} - {file_jour} - {file_periode}")
        subtitle_cell.font = Font(size=12, bold=True)
        row_idx += 2
        
        # Headers
        headers = ['Nom Étudiant', 'Présent', 'Signature', 'Observations']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_num, value=header)
            cell.font = header_style['font']
            cell.fill = header_style['fill']
            cell.alignment = header_style['alignment']
            cell.border = header_style['border']
        row_idx += 1
        
        # Rows
        for student in sorted(students):
            ws.cell(row=row_idx, column=1, value=student).border = cell_style['border']
            ws.cell(row=row_idx, column=2).border = cell_style['border']
            ws.cell(row=row_idx, column=3).border = cell_style['border']
            ws.cell(row=row_idx, column=4).border = cell_style['border']
            row_idx += 1
            
        try:
            wb.save(output_excel_path)
            print(f"Created: {output_excel_path}")
        except Exception as e:
            print(f"Error saving Excel {output_excel_path}: {e}")

    if not found_any:
        print("No matching files found.")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate attendance Excel files.")
    parser.add_argument("--semaine", type=int, help="Target week number")
    parser.add_argument("--jour", type=str, help="Target day")
    parser.add_argument("--periode", type=str, help="Target period (optional)")
    parser.add_argument("--discipline", type=str, help="Target discipline (optional)")
    parser.add_argument("--input_dir", type=str, help="Input directory containing CSVs (for batch mode)")
    parser.add_argument("--output_dir", type=str, help="Output directory")
    parser.add_argument("--batch", action="store_true", help="Run in batch mode from individual CSVs (fiches_dir)")
    parser.add_argument("--mode", choices=['single', 'batch'], default='single', help="Operation mode: single (from global planning) or batch (from fiches_dir)")

    args = parser.parse_args()
    
    base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
    
    # Defaults
    if args.batch or args.input_dir:
        # Batch mode
        input_dir = args.input_dir if args.input_dir else os.path.join(base_dir, "resultat", "fiche_appel")
        if not os.path.isabs(input_dir):
            input_dir = os.path.join(base_dir, input_dir)
            
        if not args.semaine or not args.jour:
            print("Error: --semaine and --jour are required for batch mode.")
        else:
            if args.output_dir:
                output_dir = args.output_dir
            else:
                output_dir = os.path.join(base_dir, "resultat", f"feuilles_appel_excel_S{args.semaine}_{args.jour}")
            
            if not os.path.isabs(output_dir):
                output_dir = os.path.join(base_dir, output_dir)
                
            generate_all_discipline_appel_excels_for_one_day(input_dir, output_dir, args.semaine, args.jour, args.periode)
            
    else:
        # Single file mode (from global solution)
        input_csv = os.path.join(base_dir, 'resultat', 'planning_solution.csv')
        
        # Build output filename
        suffix_parts = []
        if args.semaine: suffix_parts.append(f"S{args.semaine}")
        if args.jour: suffix_parts.append(f"{args.jour}")
        if args.periode: suffix_parts.append(f"{args.periode}")
        if args.discipline: suffix_parts.append(f"{args.discipline}")
        
        suffix = "_" + "_".join(suffix_parts) if suffix_parts else ""
        output_name = f'feuilles_appel{suffix}.xlsx'
        
        if args.output_dir:
            output_dir = args.output_dir
        else:
            output_dir = os.path.join(base_dir, 'resultat')
            
        if not os.path.isabs(output_dir):
            output_dir = os.path.join(base_dir, output_dir)
            
        output_excel = os.path.join(output_dir, output_name)
        
        generate_appel_excel(input_csv, output_excel, args.semaine, args.jour, args.periode, args.discipline)
