import csv
import os
import argparse
import datetime
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Define short codes for disciplines (Add mappings as needed)
DISC_CODES = {
    "Polyclinique": "POLY",
    "Parodontologie": "PARO",
    "Comodulation": "COMO",
    "Pédodontie Soins": "PEDO_S",
    "Orthodontie": "ODF",
    "Occlusodontie": "OCCL",
    "Radiologie": "RADIO",
    "Stérilisation": "STERI",
    "Panoramique": "PANO",
    "Urgence": "URG",
    "Pédodontie Urgences": "PEDO_U",
    "BLOC": "BLOC",
    "Soins spécifiques": "SOINS_SPE"
}

def get_short_name(full_name):
    return DISC_CODES.get(full_name, full_name[:4].upper())

def create_header_style(color="808080"):
    return {
        'font': Font(bold=True, color="FFFFFF"),
        'fill': PatternFill(start_color=color, end_color=color, fill_type="solid"),
        'alignment': Alignment(horizontal="center", vertical="center"),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    }

def create_cell_style():
    return {
        'alignment': Alignment(vertical="center", wrap_text=False),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')),
        'font': Font(size=10)
    }

def generate_discipline_year_excel(input_path: str, output_dir: str):
    """
    Generates one Excel file per discipline, formatted as a weekly schedule grid 
    (Matin/Apres-Midi sections, 5 days side-by-side).
    """
    print(f"Generating Year Recap by Discipline from: {input_path}")
    
    if not os.path.exists(input_path):
        print(f"Error: Input file {input_path} not found.")
        return

    # Data structure: discipline -> week -> period (0=AM,1=PM) -> day_idx (0-4) -> list of students
    discipline_data = {} 

    try:
        with open(input_path, mode='r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                if "STAGE:" in row["Discipline"]:
                    continue

                try:
                    semaine = int(row["Semaine"])
                    jour = row["Jour"]
                    periode = row["Apres-Midi"] 
                    discipline = row["Discipline"]
                    eleve = row["Id_Eleve"]
                except ValueError:
                    continue

                if discipline not in discipline_data:
                    discipline_data[discipline] = {}
                
                if semaine not in discipline_data[discipline]:
                    discipline_data[discipline][semaine] = {0: {d:[] for d in range(5)}, 1: {d:[] for d in range(5)}}

                day_map = {"Lundi": 0, "Mardi": 1, "Mercredi": 2, "Jeudi": 3, "Vendredi": 4}
                d_idx = day_map.get(jour, 99)
                if d_idx > 4: continue

                p_int = int(periode) if periode.isdigit() else (0 if "Matin" in periode else 1)
                
                discipline_data[discipline][semaine][p_int][d_idx].append(eleve)

    except Exception as e:
        print(f"Error reading CSV: {e}")
        return

    # Ensure output dir
    os.makedirs(output_dir, exist_ok=True)

    # Styles
    style_matin = create_header_style("660066") # Dark Purple
    style_aprem = create_header_style("990000") # Dark Red
    style_day_header = {
        'font': Font(bold=True),
        'alignment': Alignment(horizontal="center", vertical="center"),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    }
    style_std = create_cell_style()
    style_disc_col = {
        'alignment': Alignment(horizontal="center", vertical="center"),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')),
        'font': Font(size=9)
    }

    count = 0
    for discipline, weeks_map in discipline_data.items():
        disc_short = get_short_name(discipline)
        safe_disc = "".join([c if c.isalnum() else "_" for c in discipline])
        out_path = os.path.join(output_dir, f"Appel_Annuel_{safe_disc}.xlsx")
        
        wb = Workbook()
        wb.remove(wb.active) # Clear default
        
        # Create sheets for all 52 weeks
        for week_num in range(1, 53):
            ws = wb.create_sheet(title=f"Semaine {week_num}")
            
            # --- Column Configuration ---
            # Columns A, C, E, G, I -> Distance/Code (Width 5)
            # Columns B, D, F, H, J -> Student Name (Width 25)
            for i in range(5):
                col_code = chr(65 + (i*2))   # A, C, E, G, I
                col_name = chr(65 + (i*2) + 1) # B, D, F, H, J
                ws.column_dimensions[col_code].width = 6
                ws.column_dimensions[col_name].width = 25

            # --- ROW 1: Header Info ---
            # Top Left: Discipline Short Name
            ws.merge_cells('A1:B1')
            ws['A1'] = disc_short
            ws['A1'].font = Font(bold=True, size=12)
            ws['A1'].alignment = Alignment(horizontal="center")

            # Center Left: Week Number
            ws.merge_cells('C1:D1')
            ws['C1'] = f"Semaine {week_num}"
            ws['C1'].font = Font(bold=True, size=11)
            ws['C1'].alignment = Alignment(horizontal="center")

            # Center Right: Date Range (Empty)
            ws.merge_cells('E1:I1')
            ws['E1'] = ""
            ws['E1'].font = Font(bold=True, size=11)
            ws['E1'].alignment = Alignment(horizontal="center")

            current_row = 3

            # ================= SECTION MATIN =================
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
            cell = ws.cell(row=current_row, column=1, value="MATIN de 08h30 à 12h30")
            cell.font = style_matin['font']
            cell.fill = style_matin['fill']
            cell.alignment = style_matin['alignment']
            # Apply border to merged cells (approximated by setting style on first cell usually, but openpyxl requires iterating)
            # Simpler: just set the top left style
            current_row += 1

            # Day Headers (Lundi, Mardi...)
            days_fr = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi"]
            for d in range(5):
                c_start = 1 + (d * 2)
                c_end = c_start + 1
                cell = ws.cell(row=current_row, column=c_start, value=days_fr[d])
                ws.merge_cells(start_row=current_row, start_column=c_start, end_row=current_row, end_column=c_end)
                cell.font = style_day_header['font']
                cell.alignment = style_day_header['alignment']
                cell.border = style_day_header['border']
                # Border for second cell in merge
                ws.cell(row=current_row, column=c_end).border = style_day_header['border']
            current_row += 1

            # Data Rows (Matin)
            # Find max students in any day for Matin to determine height
            students_map_am = {}
            max_rows_am = 0
            if week_num in weeks_map:
                for d in range(5):
                    s_list = sorted(weeks_map[week_num][0][d])
                    students_map_am[d] = s_list
                    if len(s_list) > max_rows_am:
                        max_rows_am = len(s_list)
            
            # Fill AM rows
            # Ensure at least one empty row or as needed
            rows_to_print = max(max_rows_am, 1)
            for r in range(rows_to_print):
                # For each day
                for d in range(5):
                    st_list = students_map_am.get(d, [])
                    
                    disc_cell = ws.cell(row=current_row, column=1 + (d*2))
                    name_cell = ws.cell(row=current_row, column=2 + (d*2))
                    
                    disc_cell.border = style_disc_col['border']
                    disc_cell.alignment = style_disc_col['alignment']
                    disc_cell.font = style_disc_col['font']
                    
                    name_cell.border = style_std['border']
                    name_cell.alignment = style_std['alignment']
                    name_cell.font = style_std['font']

                    if r < len(st_list):
                        disc_cell.value = disc_short
                        name_cell.value = st_list[r]
                    else:
                        disc_cell.value = ""
                        name_cell.value = ""
                
                current_row += 1
            
            current_row += 2 # Spacer

            # ================= SECTION APRES-MIDI =================
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
            cell = ws.cell(row=current_row, column=1, value="APRES-MIDI de 14h00 à 18h00")
            cell.font = style_aprem['font']
            cell.fill = style_aprem['fill']
            cell.alignment = style_aprem['alignment']
            current_row += 1

            # Day Headers Repetition
            for d in range(5):
                c_start = 1 + (d * 2)
                c_end = c_start + 1
                cell = ws.cell(row=current_row, column=c_start, value=days_fr[d])
                ws.merge_cells(start_row=current_row, start_column=c_start, end_row=current_row, end_column=c_end)
                cell.font = style_day_header['font']
                cell.alignment = style_day_header['alignment']
                cell.border = style_day_header['border']
                ws.cell(row=current_row, column=c_end).border = style_day_header['border']
            current_row += 1

            # Data Rows (PM)
            students_map_pm = {}
            max_rows_pm = 0
            if week_num in weeks_map:
                for d in range(5):
                    s_list = sorted(weeks_map[week_num][1][d])
                    students_map_pm[d] = s_list
                    if len(s_list) > max_rows_pm:
                        max_rows_pm = len(s_list)

            rows_to_print_pm = max(max_rows_pm, 1)
            for r in range(rows_to_print_pm):
                for d in range(5):
                    st_list = students_map_pm.get(d, [])
                    
                    disc_cell = ws.cell(row=current_row, column=1 + (d*2))
                    name_cell = ws.cell(row=current_row, column=2 + (d*2))
                    
                    disc_cell.border = style_disc_col['border']
                    disc_cell.alignment = style_disc_col['alignment']
                    disc_cell.font = style_disc_col['font']
                    
                    name_cell.border = style_std['border']
                    name_cell.alignment = style_std['alignment']
                    name_cell.font = style_std['font']

                    if r < len(st_list):
                        disc_cell.value = disc_short
                        name_cell.value = st_list[r]
                    else:
                        disc_cell.value = ""
                        name_cell.value = ""
                
                current_row += 1

        try:
            wb.save(out_path)
            count += 1
            print(f"Generated {out_path}")
        except Exception as e:
            print(f"Error saving {out_path}: {e}")
    
    print(f"Finished generating {count} discipline files.")

if __name__ == "__main__":
    
    # Path navigation
    base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
    input_csv = os.path.join(base_dir, 'resultat', 'planning_solution.csv')
    
    # Simple argument parser for optional output dir
    parser = argparse.ArgumentParser(description="Generate annual attendance Excel files per discipline.")
    parser.add_argument("--output_dir", type=str, help="Output directory")
    args = parser.parse_args()

    output_dir = args.output_dir if args.output_dir else os.path.join(base_dir, 'resultat', 'fiche_appel_par_discipline')
    
    generate_discipline_year_excel(input_csv, output_dir)
