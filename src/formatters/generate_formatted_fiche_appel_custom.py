import csv
import os
import argparse
import datetime
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Define short codes for disciplines (Add mappings as needed)
DISC_CODES = {
    "Polyclinique": "POLY",
    "Parodontologie": "PARO",
    "Comodulation": "COMO",
    "Pédodontie Soins": "PEDO_S",
    "Orthodontie": "ODF",
    "Occlusodontie": "OCCL",
    "Radiologie": "RADIO",
    "Stérilisation": "STERI",
    "Panoramique": "PANO",
    "Urgence": "URG",
    "Pédodontie Urgences": "PEDO_U",
    "BLOC": "BLOC",
    "Soins spécifiques": "SOINS_SPE"
}

def get_short_name(full_name):
    return DISC_CODES.get(full_name, full_name[:4].upper())

def create_header_style(color="808080"):
    return {
        'font': Font(bold=True, color="FFFFFF"),
        'fill': PatternFill(start_color=color, end_color=color, fill_type="solid"),
        'alignment': Alignment(horizontal="center", vertical="center"),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    }

def create_cell_style():
    return {
        'alignment': Alignment(vertical="center", wrap_text=False),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')),
        'font': Font(size=10)
    }

def generate_discipline_year_excel(input_path: str, output_dir: str, csv_basename: str):
    """
    Generates one Excel file per discipline, formatted as a weekly schedule grid 
    (Matin/Apres-Midi sections, 5 days side-by-side).
    """
    print(f"Generating Year Recap by Discipline from: {input_path}")
    
    if not os.path.exists(input_path):
        print(f"Error: Input file {input_path} not found.")
        return 0

    # Data structure: discipline -> week -> period (0=AM,1=PM) -> day_idx (0-4) -> list of students
    discipline_data = {} 

    try:
        with open(input_path, mode='r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                if "STAGE:" in row["Discipline"]:
                    continue

                try:
                    discipline = row["Discipline"]
                    semaine = int(row["Semaine"])
                    jour = row["Jour"]
                    periode = row["Apres-Midi"]
                    eleve = row["Id_Eleve"]
                except ValueError:
                    continue

                if discipline not in discipline_data:
                    discipline_data[discipline] = {}
                
                if semaine not in discipline_data[discipline]:
                    discipline_data[discipline][semaine] = {0: [[], [], [], [], []], 1: [[], [], [], [], []]}

                day_map = {"Lundi": 0, "Mardi": 1, "Mercredi": 2, "Jeudi": 3, "Vendredi": 4}
                d_idx = day_map.get(jour, 99)
                if d_idx > 4: continue

                p_int = int(periode) if periode.isdigit() else (0 if "Matin" in periode else 1)
                
                discipline_data[discipline][semaine][p_int][d_idx].append(eleve)

    except Exception as e:
        print(f"Error reading CSV: {e}")
        return 0

    # Ensure output dir
    os.makedirs(output_dir, exist_ok=True)

    # Styles
    style_matin = create_header_style("660066") # Dark Purple
    style_aprem = create_header_style("990000") # Dark Red
    style_day_header = {
        'font': Font(bold=True),
        'alignment': Alignment(horizontal="center", vertical="center"),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    }
    style_std = create_cell_style()
    style_disc_col = {
        'alignment': Alignment(horizontal="center", vertical="center"),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')),
        'font': Font(size=9)
    }

    count = 0
    for discipline, weeks_map in discipline_data.items():
        disc_short = get_short_name(discipline)
        safe_disc = "".join([c if c.isalnum() else "_" for c in discipline])
        
        # Inclure le nom du CSV dans le nom du fichier
        out_filename = f"Appel_Annuel_{safe_disc}_{csv_basename}.xlsx"
        out_path = os.path.join(output_dir, out_filename)
        
        wb = Workbook()
        wb.remove(wb.active) # Clear default
        
        # Create sheets for all 52 weeks
        for week_num in range(1, 53):
            ws = wb.create_sheet(title=f"Semaine {week_num}")
            
            # --- Column Configuration ---
            # Columns A, C, E, G, I -> Distance/Code (Width 5)
            # Columns B, D, F, H, J -> Student Name (Width 25)
            for i in range(5):
                col_code = chr(65 + (i*2))   # A, C, E, G, I
                col_name = chr(65 + (i*2) + 1) # B, D, F, H, J
                ws.column_dimensions[col_code].width = 6
                ws.column_dimensions[col_name].width = 25

            # --- ROW 1: Header Info ---
            # Top Left: Discipline Short Name
            ws.merge_cells('A1:B1')
            ws['A1'] = disc_short
            ws['A1'].font = Font(bold=True, size=12)
            ws['A1'].alignment = Alignment(horizontal="center")

            # Center Left: Week Number
            ws.merge_cells('C1:D1')
            ws['C1'] = f"Semaine {week_num}"
            ws['C1'].font = Font(bold=True, size=11)
            ws['C1'].alignment = Alignment(horizontal="center")

            # Center Right: Date Range (Empty)
            ws.merge_cells('E1:I1')
            ws['E1'] = ""
            ws['E1'].font = Font(bold=True, size=11)
            ws['E1'].alignment = Alignment(horizontal="center")

            current_row = 3

            # ================= SECTION MATIN =================
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
            cell = ws.cell(row=current_row, column=1, value="MATIN de 08h30 à 12h30")
            cell.font = style_matin['font']
            cell.fill = style_matin['fill']
            cell.alignment = style_matin['alignment']
            current_row += 1

            # Day Headers (Lundi, Mardi...)
            days_fr = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi"]
            for d in range(5):
                c_start = 1 + (d * 2)
                c_end = c_start + 1
                cell = ws.cell(row=current_row, column=c_start, value=days_fr[d])
                ws.merge_cells(start_row=current_row, start_column=c_start, end_row=current_row, end_column=c_end)
                cell.font = style_day_header['font']
                cell.alignment = style_day_header['alignment']
                cell.border = style_day_header['border']
                ws.cell(row=current_row, column=c_end).border = style_day_header['border']
            current_row += 1

            # Data Rows (Matin)
            students_map_am = {}
            max_rows_am = 0
            if week_num in weeks_map:
                for d_idx in range(5):
                    students_map_am[d_idx] = weeks_map[week_num][0][d_idx]
                    max_rows_am = max(max_rows_am, len(students_map_am[d_idx]))
            
            rows_to_print = max(max_rows_am, 1)
            for r in range(rows_to_print):
                for d_idx in range(5):
                    c_start = 1 + (d_idx * 2)
                    c_end = c_start + 1
                    
                    students = students_map_am.get(d_idx, [])
                    if r < len(students):
                        student_id = students[r]
                        ws.cell(row=current_row, column=c_start, value="")
                        ws.cell(row=current_row, column=c_end, value=student_id)
                    else:
                        ws.cell(row=current_row, column=c_start, value="")
                        ws.cell(row=current_row, column=c_end, value="")
                    
                    ws.cell(row=current_row, column=c_start).border = style_std['border']
                    ws.cell(row=current_row, column=c_end).border = style_std['border']
                    ws.cell(row=current_row, column=c_start).alignment = style_std['alignment']
                    ws.cell(row=current_row, column=c_end).alignment = style_std['alignment']
                
                current_row += 1
            
            current_row += 2 # Spacer

            # ================= SECTION APRES-MIDI =================
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
            cell = ws.cell(row=current_row, column=1, value="APRES-MIDI de 14h00 à 18h00")
            cell.font = style_aprem['font']
            cell.fill = style_aprem['fill']
            cell.alignment = style_aprem['alignment']
            current_row += 1

            # Day Headers Repetition
            for d in range(5):
                c_start = 1 + (d * 2)
                c_end = c_start + 1
                cell = ws.cell(row=current_row, column=c_start, value=days_fr[d])
                ws.merge_cells(start_row=current_row, start_column=c_start, end_row=current_row, end_column=c_end)
                cell.font = style_day_header['font']
                cell.alignment = style_day_header['alignment']
                cell.border = style_day_header['border']
                ws.cell(row=current_row, column=c_end).border = style_day_header['border']
            current_row += 1

            # Data Rows (PM)
            students_map_pm = {}
            max_rows_pm = 0
            if week_num in weeks_map:
                for d_idx in range(5):
                    students_map_pm[d_idx] = weeks_map[week_num][1][d_idx]
                    max_rows_pm = max(max_rows_pm, len(students_map_pm[d_idx]))

            rows_to_print_pm = max(max_rows_pm, 1)
            for r in range(rows_to_print_pm):
                for d_idx in range(5):
                    c_start = 1 + (d_idx * 2)
                    c_end = c_start + 1
                    
                    students = students_map_pm.get(d_idx, [])
                    if r < len(students):
                        student_id = students[r]
                        ws.cell(row=current_row, column=c_start, value="")
                        ws.cell(row=current_row, column=c_end, value=student_id)
                    else:
                        ws.cell(row=current_row, column=c_start, value="")
                        ws.cell(row=current_row, column=c_end, value="")
                    
                    ws.cell(row=current_row, column=c_start).border = style_std['border']
                    ws.cell(row=current_row, column=c_end).border = style_std['border']
                    ws.cell(row=current_row, column=c_start).alignment = style_std['alignment']
                    ws.cell(row=current_row, column=c_end).alignment = style_std['alignment']
                
                current_row += 1

        try:
            wb.save(out_path)
            count += 1
            print(f"  ✓ Generated {out_filename}")
        except Exception as e:
            print(f"  ✗ Error saving {out_path}: {e}")
    
    print(f"\n✓ Finished generating {count} discipline files.")
    return count

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate annual attendance Excel files per discipline.")
    parser.add_argument("--input", type=str, required=True, help="Path to input CSV file (planning_solution.csv)")
    parser.add_argument("--output_dir", type=str, required=True, help="Output directory for fiche d'appel files")
    
    args = parser.parse_args()
    
    # Vérifier que le fichier d'entrée existe
    if not os.path.exists(args.input):
        print(f"Error: Input file not found: {args.input}")
        exit(1)
    
    # Extraire le nom de base du fichier CSV (sans extension)
    csv_basename = os.path.splitext(os.path.basename(args.input))[0]
    
    print("="*70)
    print("GÉNÉRATION DES FICHES D'APPEL PAR DISCIPLINE")
    print("="*70)
    
    count = generate_discipline_year_excel(args.input, args.output_dir, csv_basename)
    
    if count > 0:
        print("\n" + "="*70)
        print("✓ TRAITEMENT TERMINÉ")
        print("="*70)
        print(f"Fichiers générés: {args.output_dir}")
        print(f"Nombre de disciplines: {count}")
    else:
        print("\n✗ Aucune fiche d'appel générée. Vérifiez le fichier d'entrée.")
