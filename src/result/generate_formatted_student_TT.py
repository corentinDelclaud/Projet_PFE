import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime, timedelta

def create_timetable_excel(csv_folder, output_excel):
    """
    Convertit des fichiers CSV d'emplois du temps en un fichier Excel formaté.
    
    Args:
        csv_folder: Dossier contenant les fichiers CSV des élèves
        output_excel: Nom du fichier Excel de sortie
    """
    
    # Mapper les jours et périodes aux vacations 1-10
    jour_periode_to_vacation = {
        ('Lundi', 'Matin'): 1,
        ('Lundi', 'Apres-Midi'): 2,
        ('Mardi', 'Matin'): 3,
        ('Mardi', 'Apres-Midi'): 4,
        ('Mercredi', 'Matin'): 5,
        ('Mercredi', 'Apres-Midi'): 6,
        ('Jeudi', 'Matin'): 7,
        ('Jeudi', 'Apres-Midi'): 8,
        ('Vendredi', 'Matin'): 9,
        ('Vendredi', 'Apres-Midi'): 10
    }
    
    # Créer le workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    csv_files = list(Path(csv_folder).glob("*.csv"))
    print(f"Traitement de {len(csv_files)} fichiers CSV...")
    
    for idx, csv_file in enumerate(csv_files, 1):
        print(f"  [{idx}/{len(csv_files)}] Traitement de {csv_file.name}")
        
        # Lire le CSV
        df = pd.read_csv(csv_file)
        
        # Créer une nouvelle feuille avec le nom de l'élève
        student_name = df['Eleve'].iloc[0] if 'Eleve' in df.columns else csv_file.stem
        ws = wb.create_sheet(title=student_name[:31])  # Excel limite à 31 caractères
        
        # Créer l'en-tête
        headers = ['SEMAINE', 'DATE']
        jours = ['LUNDI', 'MARDI', 'MERCREDI', 'JEUDI', 'VENDREDI']
        for jour in jours:
            headers.extend([jour, ''])  # Deux colonnes par jour (vacation 1-2, 3-4, etc.)
        
        # Sous-en-têtes avec les numéros de vacation
        sub_headers = ['', '']  # Pour SEMAINE et DATE
        vacation_num = 1
        for _ in jours:
            sub_headers.extend([str(vacation_num), str(vacation_num + 1)])
            vacation_num += 2
        
        # Écrire les en-têtes
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Écrire les sous-en-têtes
        for col_idx, sub_header in enumerate(sub_headers, 1):
            cell = ws.cell(row=2, column=col_idx, value=sub_header)
            cell.font = Font(bold=True, size=10)
            cell.fill = PatternFill(start_color="7BA0C0", end_color="7BA0C0", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Fusionner les cellules d'en-tête pour chaque jour
        ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=4)   # LUNDI
        ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=6)   # MARDI
        ws.merge_cells(start_row=1, start_column=7, end_row=1, end_column=8)   # MERCREDI
        ws.merge_cells(start_row=1, start_column=9, end_row=1, end_column=10)  # JEUDI
        ws.merge_cells(start_row=1, start_column=11, end_row=1, end_column=12) # VENDREDI
        
        # Définir les largeurs de colonnes
        ws.column_dimensions['A'].width = 10  # SEMAINE
        ws.column_dimensions['B'].width = 12  # DATE
        for col in range(3, 13):  # Colonnes des vacations
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Organiser les données par semaine
        df['vacation'] = df.apply(
            lambda row: jour_periode_to_vacation.get((row['Jour'], row['Periode']), None),
            axis=1
        )
        
        # Créer un dictionnaire pour accès rapide
        planning_dict = {}
        for _, row in df.iterrows():
            semaine = int(row['Semaine'])
            vacation = row['vacation']
            discipline = row['Discipline']
            
            if semaine not in planning_dict:
                planning_dict[semaine] = {}
            
            planning_dict[semaine][vacation] = discipline
        
        # Date de début (1er septembre 2025 - semaine 36)
        start_date = datetime(2025, 9, 1)
        
        # Remplir les données pour les 53 semaines
        row_idx = 3
        for semaine in range(36, 53):  # Semaines 36-52
            date_semaine = start_date + timedelta(weeks=(semaine - 36))
            
            # Écrire le numéro de semaine et la date
            ws.cell(row=row_idx, column=1, value=semaine)
            ws.cell(row=row_idx, column=2, value=date_semaine.strftime('%d/%m/%Y'))
            
            # Remplir les vacations
            for vacation in range(1, 11):
                col_idx = 2 + vacation  # Colonne commence à 3
                discipline = planning_dict.get(semaine, {}).get(vacation, '')
                
                cell = ws.cell(row=row_idx, column=col_idx, value=discipline)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Colorer les cellules de stage
                if 'STAGE' in discipline.upper():
                    cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
            
            row_idx += 1
        
        # Semaines 1-35
        for semaine in range(1, 36):
            date_semaine = start_date + timedelta(weeks=(semaine - 36 + 52))
            
            # Écrire le numéro de semaine et la date
            ws.cell(row=row_idx, column=1, value=semaine)
            ws.cell(row=row_idx, column=2, value=date_semaine.strftime('%d/%m/%Y'))
            
            # Remplir les vacations
            for vacation in range(1, 11):
                col_idx = 2 + vacation
                discipline = planning_dict.get(semaine, {}).get(vacation, '')
                
                cell = ws.cell(row=row_idx, column=col_idx, value=discipline)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Colorer les cellules de stage
                if 'STAGE' in discipline.upper():
                    cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
            
            row_idx += 1
        
        # Ajouter les statistiques
        row_idx += 2
        ws.cell(row=row_idx, column=1, value="Statistiques par discipline:").font = Font(bold=True)
        row_idx += 1
        
        # Compter les occurrences de chaque discipline
        discipline_counts = df[~df['Discipline'].str.contains('STAGE', na=False)]['Discipline'].value_counts()
        
        for discipline, count in discipline_counts.items():
            ws.cell(row=row_idx, column=1, value=discipline)
            ws.cell(row=row_idx, column=2, value=count)
            row_idx += 1
    
    # Créer un dossier spécifique pour le fichier Excel
    excel_folder = Path('/Users/poomedy/Desktop/semester 9/PFE/Projet_PFE/resultat') / "emplois_du_temps_excel"
    excel_folder.mkdir(exist_ok=True)
    
    # Sauvegarder le fichier Excel dans ce dossier
    output_path = excel_folder / output_excel
    wb.save(output_path)
    print(f"\n✓ Fichier Excel créé: {output_path}")
    print(f"  {len(csv_files)} emplois du temps générés")


# Utilisation
if __name__ == "__main__":
    csv_folder = "/Users/poomedy/Desktop/semester 9/PFE/Projet_PFE/resultat/planning_personnel"
    create_timetable_excel(csv_folder, "emplois_du_temps.xlsx")